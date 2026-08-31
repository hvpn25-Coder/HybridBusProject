from __future__ import annotations

import math
import os
from pathlib import Path
from xml.sax.saxutils import escape

import pandas as pd
from openpyxl import load_workbook
from PIL import Image as PILImage

from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
from reportlab.lib.fonts import addMapping
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, StyleSheet1
from reportlab.lib.units import cm, mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    HRFlowable,
    Image,
    KeepTogether,
    ListFlowable,
    ListItem,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.platypus.tableofcontents import TableOfContents
from reportlab.graphics.shapes import Circle, Drawing, Line, Polygon, PolyLine, Rect, String


ROOT = Path(r"C:\TempData\Hybrid_Vehicle\HybridBusProject")
ASSET = ROOT / "tmp" / "pdfs" / "project_textbook"
OUTPUT = ROOT / "output" / "pdf" / "Hybrid_Electric_Bus_Project_Textbook.pdf"
OUTPUT.parent.mkdir(parents=True, exist_ok=True)

NAVY = HexColor("#10263D")
NAVY_2 = HexColor("#173B59")
TEAL = HexColor("#0A7F82")
TEAL_LIGHT = HexColor("#DDEFF0")
BLUE = HexColor("#2E6F9E")
SKY = HexColor("#E7F1F8")
ORANGE = HexColor("#E8873A")
GOLD = HexColor("#D5A632")
GREEN = HexColor("#2F7D5C")
RED = HexColor("#B84A4A")
DUMP = HexColor("#C24A20")
INK = HexColor("#26333D")
MID = HexColor("#66737D")
LIGHT = HexColor("#F3F6F8")
GRID = HexColor("#D8E0E5")
WHITE = colors.white


def register_fonts():
    fonts = {
        "Arial": r"C:\Windows\Fonts\arial.ttf",
        "Arial-Bold": r"C:\Windows\Fonts\arialbd.ttf",
        "Arial-Italic": r"C:\Windows\Fonts\ariali.ttf",
        "Arial-BoldItalic": r"C:\Windows\Fonts\arialbi.ttf",
        "Times": r"C:\Windows\Fonts\times.ttf",
        "Times-Bold": r"C:\Windows\Fonts\timesbd.ttf",
        "Times-Italic": r"C:\Windows\Fonts\timesi.ttf",
        "Times-BoldItalic": r"C:\Windows\Fonts\timesbi.ttf",
        "Consolas": r"C:\Windows\Fonts\consola.ttf",
        "Consolas-Bold": r"C:\Windows\Fonts\consolab.ttf",
    }
    for name, path in fonts.items():
        pdfmetrics.registerFont(TTFont(name, path))
    addMapping("Arial", 0, 0, "Arial")
    addMapping("Arial", 1, 0, "Arial-Bold")
    addMapping("Arial", 0, 1, "Arial-Italic")
    addMapping("Arial", 1, 1, "Arial-BoldItalic")
    addMapping("Times", 0, 0, "Times")
    addMapping("Times", 1, 0, "Times-Bold")
    addMapping("Times", 0, 1, "Times-Italic")
    addMapping("Times", 1, 1, "Times-BoldItalic")


register_fonts()


styles = StyleSheet1()
styles.add(ParagraphStyle(
    "CoverKicker", fontName="Arial-Bold", fontSize=10, leading=13,
    textColor=HexColor("#9ED9D9"), spaceAfter=18, tracking=1.2,
))
styles.add(ParagraphStyle(
    "CoverTitle", fontName="Times-Bold", fontSize=31, leading=34,
    textColor=WHITE, spaceAfter=14,
))
styles.add(ParagraphStyle(
    "CoverSubtitle", fontName="Arial", fontSize=14, leading=20,
    textColor=HexColor("#D9E7EF"), spaceAfter=26,
))
styles.add(ParagraphStyle(
    "CoverMeta", fontName="Arial", fontSize=9.5, leading=15,
    textColor=HexColor("#BDD0DC"),
))
styles.add(ParagraphStyle(
    "FrontTitle", fontName="Times-Bold", fontSize=21, leading=25,
    textColor=NAVY, spaceBefore=4, spaceAfter=12,
))
styles.add(ParagraphStyle(
    "FrontSubhead", fontName="Arial-Bold", fontSize=13, leading=16,
    textColor=TEAL, spaceBefore=11, spaceAfter=5, keepWithNext=True,
))
styles.add(ParagraphStyle(
    "Heading1", fontName="Times-Bold", fontSize=22, leading=25,
    textColor=NAVY, spaceBefore=4, spaceAfter=12, keepWithNext=True,
))
styles.add(ParagraphStyle(
    "Heading2", fontName="Arial-Bold", fontSize=13, leading=16,
    textColor=TEAL, spaceBefore=11, spaceAfter=5, keepWithNext=True,
))
styles.add(ParagraphStyle(
    "Heading3", fontName="Arial-Bold", fontSize=10.2, leading=13,
    textColor=NAVY_2, spaceBefore=8, spaceAfter=3, keepWithNext=True,
))
styles.add(ParagraphStyle(
    "BodyTextBook", fontName="Times", fontSize=9.6, leading=13.7,
    textColor=INK, alignment=TA_JUSTIFY, spaceAfter=7,
))
styles.add(ParagraphStyle(
    "Lead", parent=styles["BodyTextBook"], fontName="Times-Bold",
    fontSize=10.3, leading=14.5, textColor=NAVY, spaceAfter=9,
))
styles.add(ParagraphStyle(
    "Small", fontName="Arial", fontSize=7.6, leading=10.1,
    textColor=MID, spaceAfter=4,
))
styles.add(ParagraphStyle(
    "Caption", fontName="Arial", fontSize=7.8, leading=10.2,
    textColor=INK, alignment=TA_CENTER, spaceBefore=4, spaceAfter=8,
    keepWithNext=False,
))
styles.add(ParagraphStyle(
    "FigureCaption", parent=styles["Caption"],
))
styles.add(ParagraphStyle(
    "TableCaption", fontName="Arial-Bold", fontSize=8, leading=10.5,
    textColor=NAVY, alignment=TA_LEFT, spaceBefore=7, spaceAfter=4,
    keepWithNext=True,
))
styles.add(ParagraphStyle(
    "Equation", fontName="Arial", fontSize=10.2, leading=15,
    textColor=NAVY, alignment=TA_CENTER,
))
styles.add(ParagraphStyle(
    "EquationNumber", fontName="Arial", fontSize=8.8, leading=12,
    textColor=MID, alignment=TA_RIGHT,
))
styles.add(ParagraphStyle(
    "Code", fontName="Consolas", fontSize=7.6, leading=10,
    textColor=HexColor("#17324A"), leftIndent=6, rightIndent=6,
))
styles.add(ParagraphStyle(
    "CalloutTitle", fontName="Arial-Bold", fontSize=9.2, leading=12,
    textColor=NAVY, spaceAfter=3,
))
styles.add(ParagraphStyle(
    "CalloutBody", fontName="Arial", fontSize=8.5, leading=11.5,
    textColor=INK,
))
styles.add(ParagraphStyle(
    "TOCHeading", fontName="Times-Bold", fontSize=20, leading=24,
    textColor=NAVY, spaceAfter=11,
))
styles.add(ParagraphStyle(
    "Reference", fontName="Arial", fontSize=7.8, leading=10.5,
    textColor=INK, leftIndent=12, firstLineIndent=-12, spaceAfter=5,
))


class ListOfFigures(TableOfContents):
    def notify(self, kind, stuff):
        if kind == "FigureEntry":
            self.addEntry(*stuff)


class ListOfTables(TableOfContents):
    def notify(self, kind, stuff):
        if kind == "TableEntry":
            self.addEntry(*stuff)


class TextbookDocTemplate(BaseDocTemplate):
    def __init__(self, filename, **kwargs):
        super().__init__(filename, **kwargs)
        self.heading_id = 0
        frame = Frame(
            19 * mm, 17 * mm, A4[0] - 38 * mm, A4[1] - 32 * mm,
            leftPadding=0, rightPadding=0, topPadding=11 * mm, bottomPadding=8 * mm,
            id="body",
        )
        self.addPageTemplates([PageTemplate(id="textbook", frames=[frame], onPage=self.draw_page)])

    def beforeDocument(self):
        # multiBuild may lay the story out several times while resolving the
        # contents pages; stable bookmark identifiers are required each pass.
        self.heading_id = 0

    def draw_page(self, canv: canvas.Canvas, doc):
        width, height = A4
        if doc.page == 1:
            canv.saveState()
            canv.setFillColor(NAVY)
            canv.rect(0, 0, width, height, fill=1, stroke=0)
            canv.setFillColor(NAVY_2)
            canv.circle(width - 10 * mm, height - 12 * mm, 58 * mm, fill=1, stroke=0)
            canv.setFillColor(TEAL)
            canv.circle(width - 12 * mm, 18 * mm, 36 * mm, fill=1, stroke=0)
            canv.setStrokeColor(HexColor("#4BA6A7"))
            canv.setLineWidth(2)
            canv.line(19 * mm, 37 * mm, 67 * mm, 37 * mm)
            canv.restoreState()
            return
        canv.saveState()
        canv.setStrokeColor(GRID)
        canv.setLineWidth(0.5)
        canv.line(19 * mm, height - 13 * mm, width - 19 * mm, height - 13 * mm)
        canv.setFont("Arial", 7.3)
        canv.setFillColor(MID)
        canv.drawString(19 * mm, height - 10 * mm, "HYBRID-ELECTRIC BUS SYSTEM MODEL")
        canv.drawRightString(width - 19 * mm, height - 10 * mm, "Graduate Study Edition")
        canv.line(19 * mm, 13 * mm, width - 19 * mm, 13 * mm)
        canv.drawString(19 * mm, 8.5 * mm, "MATLAB / Simulink R2025a")
        canv.drawRightString(width - 19 * mm, 8.5 * mm, f"Page {doc.page}")
        canv.restoreState()

    def afterFlowable(self, flowable):
        if not isinstance(flowable, Paragraph):
            return
        style_name = flowable.style.name
        text = flowable.getPlainText()
        if style_name in ("Heading1", "Heading2", "Heading3"):
            level = {"Heading1": 0, "Heading2": 1, "Heading3": 2}[style_name]
            self.heading_id += 1
            key = f"heading_{self.heading_id}"
            self.canv.bookmarkPage(key)
            self.canv.addOutlineEntry(text, key, level=level, closed=level > 0)
            self.notify("TOCEntry", (level, text, self.page, key))
        elif style_name == "FigureCaption":
            self.notify("FigureEntry", (0, text, self.page))
        elif style_name == "TableCaption":
            self.notify("TableEntry", (0, text, self.page))


def P(text, style="BodyTextBook"):
    return Paragraph(text, styles[style])


def bullet_list(items, level=0):
    return ListFlowable(
        [ListItem(P(item, "BodyTextBook"), leftIndent=12) for item in items],
        bulletType="bullet", start="circle", leftIndent=15 + 10 * level,
        bulletFontName="Arial", bulletFontSize=6.5, bulletColor=TEAL,
        spaceAfter=6,
    )


def numbered_list(items):
    return ListFlowable(
        [ListItem(P(item, "BodyTextBook"), leftIndent=14) for item in items],
        bulletType="1", start="1", leftIndent=19, bulletFontName="Arial-Bold",
        bulletFontSize=8, bulletColor=TEAL, spaceAfter=6,
    )


def callout(title, body, tone="info"):
    palette = {
        "info": (SKY, BLUE),
        "note": (TEAL_LIGHT, TEAL),
        "warn": (HexColor("#FFF2E6"), ORANGE),
        "critical": (HexColor("#F8E7E7"), RED),
    }
    fill, edge = palette[tone]
    content = P(f"<b>{escape(title)}</b><br/>{body}", "CalloutBody")
    table = Table([[content]], colWidths=[170 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), fill),
        ("BOX", (0, 0), (-1, -1), 0.9, edge),
        ("LEFTPADDING", (0, 0), (-1, -1), 9),
        ("RIGHTPADDING", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    return KeepTogether([Spacer(1, 3), table, Spacer(1, 7)])


def equation(formula, number):
    table = Table(
        [[P(formula, "Equation"), P(f"({number})", "EquationNumber")]],
        colWidths=[151 * mm, 19 * mm],
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), HexColor("#F6F9FB")),
        ("BOX", (0, 0), (-1, -1), 0.5, GRID),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return KeepTogether([Spacer(1, 4), table, Spacer(1, 8)])


def code_block(code):
    text = "<br/>".join(escape(line).replace(" ", "&nbsp;") for line in code.strip().splitlines())
    table = Table([[P(text, "Code")]], colWidths=[170 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), HexColor("#EEF3F6")),
        ("BOX", (0, 0), (-1, -1), 0.6, HexColor("#B7C7D1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    return KeepTogether([table, Spacer(1, 7)])


def table_caption(number, title):
    return P(f"Table {number}. {escape(title)}", "TableCaption")


def figure_caption(number, title):
    return P(f"Figure {number}. {escape(title)}", "FigureCaption")


def make_table(rows, widths, header=True, font_size=7.2, alignments=None):
    cooked = []
    for r, row in enumerate(rows):
        cooked.append([
            cell if isinstance(cell, Paragraph) else P(escape(str(cell)), "Small")
            for cell in row
        ])
    table = Table(cooked, colWidths=widths, repeatRows=1 if header else 0, hAlign="LEFT")
    commands = [
        ("BOX", (0, 0), (-1, -1), 0.6, HexColor("#AFC0CA")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, HexColor("#D4DEE4")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("FONTNAME", (0, 0), (-1, -1), "Arial"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("TEXTCOLOR", (0, 0), (-1, -1), INK),
    ]
    if header:
        commands += [
            ("BACKGROUND", (0, 0), (-1, 0), NAVY_2),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Arial-Bold"),
            ("TOPPADDING", (0, 0), (-1, 0), 5),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
        ]
        for r in range(1, len(rows)):
            if r % 2 == 0:
                commands.append(("BACKGROUND", (0, r), (-1, r), HexColor("#F6F8FA")))
    if alignments:
        for col, alignment in enumerate(alignments):
            commands.append(("ALIGN", (col, 1 if header else 0), (col, -1), alignment))
    table.setStyle(TableStyle(commands))
    return table


def arrow(d, x1, y1, x2, y2, color=INK, width=1.4):
    d.add(Line(x1, y1, x2, y2, strokeColor=color, strokeWidth=width))
    angle = math.atan2(y2 - y1, x2 - x1)
    size = 5
    pts = [
        x2, y2,
        x2 - size * math.cos(angle - 0.45), y2 - size * math.sin(angle - 0.45),
        x2 - size * math.cos(angle + 0.45), y2 - size * math.sin(angle + 0.45),
    ]
    d.add(Polygon(pts, fillColor=color, strokeColor=color))


def box(d, x, y, w, h, label, fill=WHITE, edge=TEAL, font=7.2):
    d.add(Rect(x, y, w, h, rx=5, ry=5, fillColor=fill, strokeColor=edge, strokeWidth=1.2))
    lines = label.split("\n")
    for i, line in enumerate(lines):
        d.add(String(x + w / 2, y + h / 2 + (len(lines) - 1 - 2 * i) * 4,
                     line, fontName="Arial-Bold", fontSize=font,
                     fillColor=NAVY, textAnchor="middle"))


def architecture_diagram():
    d = Drawing(485, 235)
    d.add(Rect(0, 0, 485, 235, fillColor=HexColor("#FBFCFD"), strokeColor=GRID))
    labels = [
        (10, 175, 82, 36, "Routes +\nExcel data", SKY),
        (110, 175, 88, 36, "Input validation\n+ resampling", TEAL_LIGHT),
        (216, 175, 90, 36, "Longitudinal\ndynamics", HexColor("#E9EFF7")),
        (324, 175, 75, 36, "Hub motor\ndrive", HexColor("#F5EBDD")),
        (417, 175, 58, 36, "DC bus", HexColor("#F4E6E6")),
        (220, 105, 95, 42, "Supervisory\nenergy manager", TEAL_LIGHT),
        (340, 95, 63, 42, "Battery 1", SKY),
        (412, 95, 63, 42, "Battery 2", SKY),
        (107, 95, 82, 42, "Engine\ngenset", HexColor("#FFF1E4")),
        (10, 95, 75, 42, "Auxiliary\nloads", HexColor("#F3F0E4")),
        (118, 25, 112, 38, "Energy + cost\naccounting", HexColor("#E8F3EA")),
        (262, 25, 102, 38, "MAT / CSV\nresults", HexColor("#EAF0F4")),
        (389, 25, 86, 38, "App +\noptimizer", HexColor("#EFE9F6")),
    ]
    for x, y, w, h, label, fill in labels:
        box(d, x, y, w, h, label, fill=fill)
    for a, b in [((92, 193), (110, 193)), ((198, 193), (216, 193)), ((306, 193), (324, 193)), ((399, 193), (417, 193))]:
        arrow(d, *a, *b, TEAL)
    arrow(d, 446, 175, 446, 137, RED)
    arrow(d, 417, 116, 403, 116, BLUE)
    arrow(d, 340, 116, 315, 126, BLUE)
    arrow(d, 189, 116, 220, 126, ORANGE)
    arrow(d, 85, 116, 220, 116, GOLD)
    arrow(d, 267, 105, 180, 63, GREEN)
    arrow(d, 230, 44, 262, 44, GREEN)
    arrow(d, 364, 44, 389, 44, TEAL)
    return d


def energy_flow_diagram():
    d = Drawing(485, 215)
    d.add(Rect(0, 0, 485, 215, fillColor=WHITE, strokeColor=GRID))
    box(d, 10, 158, 58, 34, "Diesel\nfuel", fill=HexColor("#FFF0E1"), edge=ORANGE)
    box(d, 86, 158, 82, 34, "Engine +\ngenerator", fill=HexColor("#FFF0E1"), edge=ORANGE)
    box(d, 188, 158, 94, 34, "Constant-point\ncharger", fill=HexColor("#FFF0E1"), edge=ORANGE)
    box(d, 305, 158, 104, 34, "Standby charge\nselector", fill=HexColor("#FFF0E1"), edge=ORANGE)
    box(d, 105, 84, 100, 38, "Battery 1\n(active/standby)", fill=SKY, edge=BLUE)
    box(d, 235, 84, 100, 38, "Battery 2\n(active/standby)", fill=SKY, edge=BLUE)
    box(d, 175, 20, 105, 38, "2 Active battery\nselector", fill=TEAL_LIGHT, edge=TEAL)
    box(d, 305, 20, 64, 38, "Traction\nDC bus", fill=HexColor("#F5E8E8"), edge=RED)
    box(d, 394, 20, 80, 38, "Hub motors +\nwheels", fill=HexColor("#E8F3EA"), edge=GREEN)
    box(d, 10, 20, 68, 38, "1 Auxiliary\nloads", fill=HexColor("#F3F0E4"), edge=GOLD)
    box(d, 90, 20, 70, 38, "3 Resistor\nload bank", fill=HexColor("#F8E7E2"), edge=DUMP)
    arrow(d, 68, 175, 86, 175, ORANGE, 2)
    arrow(d, 168, 175, 188, 175, ORANGE, 2)
    arrow(d, 282, 175, 305, 175, ORANGE, 2)
    arrow(d, 337, 158, 180, 122, ORANGE, 1.8)
    arrow(d, 377, 158, 285, 122, ORANGE, 1.8)
    arrow(d, 155, 84, 205, 58, BLUE, 1.8)
    arrow(d, 285, 84, 250, 58, BLUE, 1.8)
    arrow(d, 280, 39, 305, 39, RED, 2)
    arrow(d, 369, 39, 394, 39, RED, 2)
    arrow(d, 394, 31, 369, 31, GREEN, 1.4)
    d.add(PolyLine([337, 58, 337, 76, 44, 76], strokeColor=GOLD, strokeWidth=1.4))
    arrow(d, 44, 76, 44, 58, GOLD, 1.4)
    d.add(PolyLine([347, 58, 347, 68, 125, 68], strokeColor=DUMP, strokeWidth=1.4))
    arrow(d, 125, 68, 125, 58, DUMP, 1.4)
    d.add(String(240, 202, "isolated standby charging path - no connection to traction bus", fontName="Arial-Bold", fontSize=7.2, fillColor=ORANGE, textAnchor="middle"))
    d.add(String(240, 7, "regeneration priority: 1 auxiliary loads  -  2 active battery  -  3 resistor load bank", fontName="Arial-Bold", fontSize=7, fillColor=MID, textAnchor="middle"))
    return d


def pipeline_diagram():
    d = Drawing(485, 145)
    d.add(Rect(0, 0, 485, 145, fillColor=HexColor("#FBFCFD"), strokeColor=GRID))
    stages = [
        (10, "Source\ntraces"), (88, "Route\nconversion"), (174, "Excel\ndatabase"),
        (260, "Schema +\nphysics checks"), (355, "Resolved\nInput struct"), (430, "1 s\nroute")
    ]
    widths = [60, 68, 68, 76, 62, 45]
    for (x, label), w in zip(stages, widths):
        box(d, x, 70, w, 38, label, fill=SKY if x < 174 else TEAL_LIGHT)
    for i in range(len(stages) - 1):
        x1 = stages[i][0] + widths[i]
        x2 = stages[i + 1][0]
        arrow(d, x1, 89, x2, 89, TEAL)
    d.add(String(10, 40, "VECTO .vdri", fontName="Arial", fontSize=7, fillColor=MID))
    d.add(String(88, 40, "1 Hz acceleration-limited", fontName="Arial", fontSize=7, fillColor=MID))
    d.add(String(260, 40, "IDs, ranges, monotonic time", fontName="Arial", fontSize=7, fillColor=MID))
    d.add(String(355, 40, "selected catalog rows", fontName="Arial", fontSize=7, fillColor=MID))
    d.add(String(10, 25, "OSM/OSRM JSON", fontName="Arial", fontSize=7, fillColor=MID))
    d.add(String(88, 25, "10 s coach adaptation", fontName="Arial", fontSize=7, fillColor=MID))
    d.add(String(430, 25, "Simulink", fontName="Arial", fontSize=7, fillColor=MID))
    return d


def supervisor_diagram():
    d = Drawing(485, 250)
    d.add(Rect(0, 0, 485, 250, fillColor=HexColor("#FBFCFD"), strokeColor=GRID))
    states = [
        (30, 175, 95, 40, "Mode 1\nB1 traction", SKY),
        (360, 175, 95, 40, "Mode 2\nB2 traction", SKY),
        (30, 85, 125, 42, "Mode 3\nB1 traction\ncharge standby B2", TEAL_LIGHT),
        (330, 85, 125, 42, "Mode 4\nB2 traction\ncharge standby B1", TEAL_LIGHT),
        (187, 155, 110, 50, "30% role-swap\nguard", HexColor("#FFF0E1")),
        (187, 70, 110, 42, "Mode 6\nRegeneration", HexColor("#E8F3EA")),
        (187, 12, 110, 34, "Mode 7\nProtection", HexColor("#F8E7E7")),
    ]
    for x, y, w, h, label, fill in states:
        box(d, x, y, w, h, label, fill=fill, edge=RED if "Protection" in label else TEAL)
    arrow(d, 125, 195, 187, 180, TEAL)
    arrow(d, 360, 195, 297, 180, TEAL)
    arrow(d, 77, 175, 77, 127, ORANGE)
    arrow(d, 408, 175, 408, 127, ORANGE)
    arrow(d, 155, 106, 187, 168, BLUE)
    arrow(d, 330, 106, 297, 168, BLUE)
    arrow(d, 242, 155, 242, 112, GREEN)
    arrow(d, 242, 70, 242, 46, RED)
    arrow(d, 187, 91, 125, 195, TEAL)
    arrow(d, 297, 91, 360, 195, TEAL)
    d.add(String(242, 232, "active SOE <= 30% swaps roles only when the alternate pack is above 30%", fontName="Arial-Bold", fontSize=7.3, fillColor=MID, textAnchor="middle"))
    d.add(String(242, 136, "genset runs at one optimum point and charges standby only", fontName="Arial", fontSize=7, fillColor=ORANGE, textAnchor="middle"))
    return d


def regeneration_priority_diagram():
    d = Drawing(485, 185)
    d.add(Rect(0, 0, 485, 185, fillColor=HexColor("#FBFCFD"), strokeColor=GRID))
    d.add(String(242, 166, "STRICT SEQUENTIAL REGENERATIVE-ENERGY ALLOCATION",
                 fontName="Arial-Bold", fontSize=8, fillColor=NAVY, textAnchor="middle"))
    box(d, 10, 105, 78, 42, "Available\nregen DC", fill=HexColor("#E8F3EA"), edge=GREEN)
    box(d, 108, 105, 92, 42, "1 Auxiliary\nloads", fill=HexColor("#F3F0E4"), edge=GOLD)
    box(d, 225, 105, 110, 42, "2 Active battery\ncharge limits", fill=SKY, edge=BLUE)
    box(d, 365, 105, 110, 42, "3 Resistor\nload bank", fill=HexColor("#F8E7E2"), edge=DUMP)
    arrow(d, 88, 126, 108, 126, GREEN, 1.8)
    arrow(d, 200, 126, 225, 126, BLUE, 1.8)
    arrow(d, 335, 126, 365, 126, DUMP, 1.8)
    d.add(String(49, 82, "P_reg = max(0, -P_motor,dc)", fontName="Arial", fontSize=6.7,
                 fillColor=MID, textAnchor="middle"))
    d.add(String(154, 82, "min(P_reg, P_aux)", fontName="Arial", fontSize=6.7,
                 fillColor=MID, textAnchor="middle"))
    d.add(String(280, 82, "remaining power within charge/SOE limits", fontName="Arial",
                 fontSize=6.7, fillColor=MID, textAnchor="middle"))
    d.add(String(420, 82, "all unaccepted remainder", fontName="Arial", fontSize=6.7,
                 fillColor=MID, textAnchor="middle"))
    d.add(Rect(55, 22, 375, 35, rx=4, ry=4, fillColor=WHITE, strokeColor=TEAL, strokeWidth=1))
    d.add(String(242, 42,
                 "P_reg = P_reg,aux + P_reg,active + P_dump",
                 fontName="Arial-Bold", fontSize=8.2, fillColor=TEAL, textAnchor="middle"))
    d.add(String(242, 29,
                 "The standby battery is never a regenerative-energy destination.",
                 fontName="Arial", fontSize=7, fillColor=MID, textAnchor="middle"))
    return d


def optimization_diagram():
    d = Drawing(485, 175)
    d.add(Rect(0, 0, 485, 175, fillColor=WHITE, strokeColor=GRID))
    labels = [
        (10, 105, 72, "Candidate\nIDs"), (101, 105, 82, "Compatibility\nfilter"),
        (202, 105, 75, "Prepare\ninputs"), (296, 105, 72, "Run\nkernel"),
        (387, 105, 88, "Feasibility\nchecks"), (202, 35, 75, "Rank by\nEUR/km"),
        (315, 35, 92, "Top 10 +\nbest result"),
    ]
    for x, y, w, label in labels:
        box(d, x, y, w, 38, label, fill=TEAL_LIGHT if y > 80 else SKY)
    for i in range(4):
        x1 = labels[i][0] + labels[i][2]
        x2 = labels[i + 1][0]
        arrow(d, x1, 124, x2, 124, TEAL)
    arrow(d, 431, 105, 277, 54, GREEN)
    arrow(d, 277, 54, 315, 54, GREEN)
    d.add(String(123, 75, "reject: voltage, speed, ratio, power", fontName="Arial", fontSize=7, fillColor=RED, textAnchor="middle"))
    d.add(String(415, 75, "reject: unmet energy, residual, terminal SOE", fontName="Arial", fontSize=7, fillColor=RED, textAnchor="middle"))
    return d


def validation_pyramid():
    d = Drawing(485, 220)
    d.add(Rect(0, 0, 485, 220, fillColor=HexColor("#FBFCFD"), strokeColor=GRID))
    levels = [
        (75, 25, 335, 42, "System / ranking: repeatability, optimization order", HexColor("#DDEFF0")),
        (115, 72, 255, 42, "Controller: switching, hysteresis, limits", HexColor("#E7F1F8")),
        (155, 119, 175, 42, "Energy: DC balance, cost correction", HexColor("#FFF0E1")),
        (195, 166, 95, 35, "Physics: force + power", HexColor("#E8F3EA")),
    ]
    for x, y, w, h, label, fill in levels:
        d.add(Polygon([x, y, x + w, y, x + w - 20, y + h, x + 20, y + h], fillColor=fill, strokeColor=TEAL))
        d.add(String(242, y + h / 2 - 2, label, fontName="Arial-Bold", fontSize=7.4, fillColor=NAVY, textAnchor="middle"))
    return d


def roadmap_diagram():
    d = Drawing(485, 190)
    d.add(Rect(0, 0, 485, 190, fillColor=WHITE, strokeColor=GRID))
    stages = [
        (15, 100, "Concept\nenergy model", GREEN),
        (113, 100, "Calibrated\ncomponent maps", TEAL),
        (211, 100, "Thermal +\nageing models", BLUE),
        (309, 100, "Closed-loop\ndriver + route", ORANGE),
        (407, 100, "HIL / fleet\nvalidation", RED),
    ]
    for x, y, label, color in stages:
        d.add(Circle(x + 30, y + 20, 28, fillColor=HexColor("#F7FAFB"), strokeColor=color, strokeWidth=2))
        for i, line in enumerate(label.split("\n")):
            d.add(String(x + 30, y + 23 - i * 9, line, fontName="Arial-Bold", fontSize=6.8, fillColor=NAVY, textAnchor="middle"))
    for i in range(len(stages) - 1):
        arrow(d, stages[i][0] + 58, 120, stages[i + 1][0] + 2, 120, MID)
    d.add(String(242, 45, "increasing fidelity, data quality, and test evidence", fontName="Arial", fontSize=8, fillColor=MID, textAnchor="middle"))
    return d


def line_chart(x, series, title, x_label, y_label, width=485, height=235, y_min=None, y_max=None):
    d = Drawing(width, height)
    d.add(Rect(0, 0, width, height, fillColor=WHITE, strokeColor=GRID))
    left, bottom, right, top = 48, 36, 18, 29
    pw, ph = width - left - right, height - bottom - top
    x = list(map(float, x))
    all_y = [float(v) for _, ys, _ in series for v in ys if math.isfinite(float(v))]
    xmin, xmax = min(x), max(x)
    ymin = min(all_y) if y_min is None else y_min
    ymax = max(all_y) if y_max is None else y_max
    if ymax <= ymin:
        ymax = ymin + 1
    pad = 0.06 * (ymax - ymin)
    if y_min is None:
        ymin -= pad
    if y_max is None:
        ymax += pad
    sx = lambda v: left + (v - xmin) / max(xmax - xmin, 1e-12) * pw
    sy = lambda v: bottom + (v - ymin) / max(ymax - ymin, 1e-12) * ph
    for i in range(6):
        yy = bottom + i * ph / 5
        d.add(Line(left, yy, left + pw, yy, strokeColor=GRID, strokeWidth=0.4))
        val = ymin + i * (ymax - ymin) / 5
        d.add(String(left - 5, yy - 2, f"{val:.1f}", fontName="Arial", fontSize=6.3, fillColor=MID, textAnchor="end"))
    for i in range(6):
        xx = left + i * pw / 5
        d.add(Line(xx, bottom, xx, bottom + ph, strokeColor=GRID, strokeWidth=0.3))
        val = xmin + i * (xmax - xmin) / 5
        d.add(String(xx, bottom - 11, f"{val:.0f}", fontName="Arial", fontSize=6.3, fillColor=MID, textAnchor="middle"))
    d.add(Line(left, bottom, left + pw, bottom, strokeColor=INK, strokeWidth=0.8))
    d.add(Line(left, bottom, left, bottom + ph, strokeColor=INK, strokeWidth=0.8))
    for label, ys, color in series:
        points = []
        for xv, yv in zip(x, ys):
            if math.isfinite(float(yv)):
                points.extend([sx(float(xv)), sy(float(yv))])
        d.add(PolyLine(points, strokeColor=color, strokeWidth=1.25))
    d.add(String(width / 2, height - 15, title, fontName="Arial-Bold", fontSize=9.2, fillColor=NAVY, textAnchor="middle"))
    d.add(String(width / 2, 8, x_label, fontName="Arial", fontSize=7, fillColor=MID, textAnchor="middle"))
    d.add(String(8, bottom + ph / 2, y_label, fontName="Arial", fontSize=7, fillColor=MID, textAnchor="middle", angle=90))
    legend_x = left + 5
    legend_y = height - 27
    for label, _, color in series:
        d.add(Line(legend_x, legend_y, legend_x + 13, legend_y, strokeColor=color, strokeWidth=2))
        d.add(String(legend_x + 17, legend_y - 2, label, fontName="Arial", fontSize=6.6, fillColor=INK))
        legend_x += 17 + pdfmetrics.stringWidth(label, "Arial", 6.6) + 19
    return d


def horizontal_bar_chart(labels, values, title, value_suffix="", width=485, height=265, max_value=None, colors_list=None):
    d = Drawing(width, height)
    d.add(Rect(0, 0, width, height, fillColor=WHITE, strokeColor=GRID))
    left, right, top, bottom = 137, 44, 30, 22
    pw, ph = width - left - right, height - top - bottom
    vmax = max(values) if max_value is None else max_value
    n = len(values)
    gap = ph / n
    for i, (label, value) in enumerate(zip(labels, values)):
        y = bottom + ph - (i + 0.72) * gap
        d.add(String(left - 7, y + 2, label, fontName="Arial", fontSize=6.8, fillColor=INK, textAnchor="end"))
        bar_w = max(0, float(value)) / max(vmax, 1e-9) * pw
        color = colors_list[i] if colors_list else TEAL
        d.add(Rect(left, y, bar_w, max(5, gap * 0.52), fillColor=color, strokeColor=None))
        d.add(String(min(left + bar_w + 4, width - right + 2), y + 2, f"{value:.1f}{value_suffix}", fontName="Arial-Bold", fontSize=6.7, fillColor=NAVY))
    d.add(String(width / 2, height - 15, title, fontName="Arial-Bold", fontSize=9.2, fillColor=NAVY, textAnchor="middle"))
    return d


def image_flow(path, max_width=170 * mm, max_height=100 * mm):
    with PILImage.open(path) as im:
        w, h = im.size
    scale = min(max_width / w, max_height / h)
    return Image(str(path), width=w * scale, height=h * scale, hAlign="CENTER")


def chapter(story, number, title):
    story.append(PageBreak())
    story.append(P(f"{number}  {title}", "Heading1"))
    story.append(HRFlowable(width="100%", thickness=1.2, color=TEAL, spaceAfter=10))


def section(story, number, title):
    story.append(P(f"{number}  {title}", "Heading2"))


def subsection(story, number, title):
    story.append(P(f"{number}  {title}", "Heading3"))


route_catalog = pd.read_csv(ASSET / "route_catalog.csv")
summary = pd.read_csv(ASSET / "default_summary.csv").iloc[0]
signals = pd.read_csv(ASSET / "default_signals.csv")
regen_priority = pd.read_csv(ASSET / "regen_priority_case.csv")
urban = pd.read_csv(ASSET / "vecto_urban_route.csv")
mass_catalog = pd.read_csv(ASSET / "mass_catalog.csv")
mass_sweep = pd.read_csv(ASSET / "mass_sweep.csv")
batteries = pd.read_csv(ASSET / "battery_catalog.csv")
motors = pd.read_csv(ASSET / "motor_catalog.csv")
gensets = pd.read_csv(ASSET / "genset_catalog.csv")
fuel_map = pd.read_csv(ASSET / "fuel_map.csv")
generator_map = pd.read_csv(ASSET / "generator_map.csv")
long_study = pd.read_csv(ASSET / "long_route_study.csv")
tests = pd.read_csv(ROOT / "results" / "HybridBus_TestResults.csv")
top_configs = pd.read_csv(ASSET / "current_top_configurations.csv")
credibility_gates = pd.read_csv(ROOT / "results" / "Model_Credibility_Summary.csv")
equivalence_checks = pd.read_csv(ROOT / "results" / "MATLAB_Simulink_Equivalence.csv")
concept_comparison = pd.read_csv(ROOT / "results" / "Powertrain_Concept_Comparison.csv")
sensitivity_study = pd.read_csv(ROOT / "results" / "Sensitivity_Study.csv")
powertrain_comparison = pd.read_csv(ASSET / "powertrain_comparison.csv")

wb = load_workbook(ROOT / "data" / "HybridBus_ComponentDatabase.xlsx", read_only=True, data_only=True)
dashboard = {}
for row in wb["Dashboard"].iter_rows(min_row=2, values_only=True):
    if row[0] is not None:
        dashboard[str(row[0])] = row[1]
sheet_names = wb.sheetnames
wb.close()


story = []

# Cover
story += [
    Spacer(1, 44 * mm),
    P("GRADUATE STUDY EDITION", "CoverKicker"),
    P("Hybrid-Electric Bus<br/>System Modeling", "CoverTitle"),
    P("A MATLAB and Simulink textbook for selectable hybrid and battery-electric<br/>architectures, route studies, supervisory control, and configuration optimization", "CoverSubtitle"),
    Spacer(1, 52 * mm),
    P(f"Project database version {dashboard.get('DatabaseVersion')}<br/>Document revision 2.3<br/>MATLAB / Simulink R2025a<br/>Revised 31 August 2026", "CoverMeta"),
    PageBreak(),
]

# Preface
story += [
    P("Preface", "FrontTitle"),
    P("This book explains a complete concept-level bus powertrain study environment with selectable hybrid-electric and battery-electric architectures. It is written for graduate students who already know basic mechanics, electrical power, and MATLAB syntax, but who may be new to model-based design and energy-management architecture.", "Lead"),
    P("The project combines a transparent backward-demand energy model, constrained-speed and forward-performance formulations, three editable Simulink representations, modular MATLAB/MAT component data, a programmatic MATLAB app, a bounded configuration optimizer, and automated verification. Its purpose is educational and architectural: it shows how equations, assumptions, data, software interfaces, and test evidence must remain consistent across a system model.", "BodyTextBook"),
    callout("How to read this book", "Chapters 1-3 establish the system and data architecture. Chapters 4-10 derive the physical and supervisory models. Chapters 11-15 explain execution, optimization, routes, and validation. Chapters 16-17 interpret worked studies and define responsible extensions.", "note"),
    P("Learning outcomes", "FrontSubhead"),
    bullet_list([
        "Derive tractive force and wheel power from a prescribed route speed and grade.",
        "Separate the traction DC-bus path from the genset-to-standby charging path.",
        "Explain a deterministic 30% dual-battery role-swap strategy and constant-point genset control.",
        "Configure Hybrid whole battery sets and BEV half-set increments, explain parallel capability-based power sharing, and distinguish charge-depleting from charge-sustaining assessment.",
        "Derive the auxiliary-first, active-battery-second, resistor-bank-third regeneration allocation and verify its power balance.",
        "Distinguish energy accounting, equivalent replenishment, charge sustaining comparison, and true feasibility.",
        "Audit route provenance, component assumptions, numerical implementation, and validation evidence.",
        "Design extensions that increase fidelity without hiding uncertainty.",
    ]),
    P("Scope statement", "FrontSubhead"),
    P("This is a concept energy model. It does not claim certification, manufacturer performance, thermal safety, battery ageing, emissions compliance, or real-time controller readiness. Synthetic component values are used deliberately so that the architecture can be studied without implying supplier data.", "BodyTextBook"),
    PageBreak(),
]

# TOC
story.append(P("Table of Contents", "TOCHeading"))
toc = TableOfContents()
toc.levelStyles = [
    ParagraphStyle("TOC1", fontName="Arial-Bold", fontSize=9.2, leading=13, leftIndent=0, firstLineIndent=0, textColor=NAVY, spaceBefore=4),
    ParagraphStyle("TOC2", fontName="Arial", fontSize=8.2, leading=11.2, leftIndent=14, firstLineIndent=0, textColor=INK),
    ParagraphStyle("TOC3", fontName="Arial", fontSize=7.6, leading=10.2, leftIndent=28, firstLineIndent=0, textColor=MID),
]
toc.dotsMinLevel = 0
story += [toc, PageBreak()]

story.append(P("Table of Figures", "TOCHeading"))
lof = ListOfFigures()
lof.levelStyles = [ParagraphStyle("LOF", fontName="Arial", fontSize=8.2, leading=11.5, leftIndent=0, textColor=INK)]
lof.dotsMinLevel = 0
story += [lof, PageBreak()]

story.append(P("List of Tables", "TOCHeading"))
lot = ListOfTables()
lot.levelStyles = [ParagraphStyle("LOT", fontName="Arial", fontSize=8.2, leading=11.5, leftIndent=0, textColor=INK)]
lot.dotsMinLevel = 0
story += [lot, PageBreak()]

# Nomenclature
story.append(P("Nomenclature and Conventions", "FrontTitle"))
nomenclature = [
    ["Symbol / term", "Meaning", "Typical unit"],
    ["v, a", "vehicle speed and longitudinal acceleration", "m/s, m/s2"],
    ["m", "calculated curb mass plus user-entered load", "kg"],
    [P("F<sub>trac</sub>", "Small"), "net tractive force required at the tyre contact patch", "N"],
    [P("P<sub>wheel</sub>", "Small"), "wheel mechanical power; positive traction, negative braking", "kW"],
    [P("P<sub>motor,dc</sub>", "Small"), "motor-side DC bus power; negative during regeneration", "kW"],
    [P("P<sub>reg</sub>", "Small"), "non-negative regenerative DC power available for allocation", "kW"],
    [P("P<sub>dump</sub>", "Small"), "surplus regenerative power dissipated by the resistor load bank", "kW"],
    ["SOE", "stored usable battery energy divided by usable-energy rating", "fraction or %"],
    ["BSFC", "brake-specific fuel consumption", "g/kWh"],
    ["η", "directional efficiency", "fraction"],
    [P("C<sub>rr</sub>", "Small"), "rolling resistance coefficient", "1"],
    ["ρ", "air density", "kg/m3"],
    ["Cd, A", "drag coefficient and frontal area", "1, m2"],
]
story += [table_caption("F.1", "Principal symbols and units"), make_table(nomenclature, [29*mm, 105*mm, 36*mm], alignments=["LEFT", "LEFT", "CENTER"])]
story.append(callout("Sign convention", "Battery power is positive when a pack discharges to the traction DC bus and negative while charging. Wheel power is positive for traction and negative for braking. Genset power is positive into the isolated standby charger; it never offsets traction demand.", "info"))

# Chapter 1
chapter(story, "1", "Project Orientation")
story.append(P("The project asks a systems question: for a prescribed bus mission, selected component set, loading state, environment, and supervisory calibration, how much wheel energy, battery energy, diesel fuel, grid-replenishment energy, operating cost, and feasible range result?", "Lead"))
section(story, "1.1", "Why multiple simulation formulations?")
story.append(P("A backward-facing model starts from the route speed that the vehicle is assumed to achieve. It differentiates speed to obtain acceleration, computes the tractive force required at the wheels, and then propagates power demand backward through the driveline to the DC bus and energy sources. This is efficient for energy sizing and comparison because it avoids solving a driver controller and vehicle-speed tracking loop.", "BodyTextBook"))
story.append(P("The app also provides a fast constrained formulation and a forward Performance formulation. Constrained mode suppresses requested acceleration using available battery current and energy, motor torque-speed/power, final-drive force, and road load while retaining the backward energy kernel. Performance mode closes the loop around desired speed, component limits, and the longitudinal vehicle plant. These modes expose achieved speed, distance completion, tracking error, limiting cause, depletion, and stall behavior.", "BodyTextBook"))
story.append(callout("Interpretation boundary", "Use Backward mode for required-energy screening, Constrained mode for fast feasibility screening, and Performance mode when achieved speed and mission completion are the decision variables. A prescribed-speed result alone is not evidence that the vehicle can track the route.", "warn"))
section(story, "1.2", "Project deliverables")
artifact_rows = [
    ["Artifact", "Role in the study"],
    ["data/HybridBus_ComponentDatabase.xlsx", "Selections, common calibration, prices, tyre, final-drive, and governance data"],
    ["data/routes/*.mat; data/batteries/*.m; data/motors/*.m; data/gensets/*.m", "Extensible one-file-per-variant route and component definitions"],
    ["models/HybridBus_BackwardModel.slx", "Editable ordinary-block hybrid backward-demand representation"],
    ["models/HybridBus_BEVModel.slx", "Separate editable BEV Simulink model with scalable parallel battery banks and no genset"],
    ["models/HybridBus_PerformanceModel.slx", "Editable forward-performance plant with route, driver, component-limit, and vehicle subsystems"],
    ["simulate_hybrid_bus_core.m", "Detailed discrete reference kernel used by batch runs and optimization"],
    ["HybridBusApp.m", "Explorer app for selecting, simulating, plotting, optimizing, and exporting"],
    ["tests/run_all_hybrid_bus_tests.m", "Fifty assertion-based physics, data, control, braking, performance, and ranking scenarios"],
    ["tests/*.m", "Fifty-six MATLAB unit and integration tests, including Hybrid/BEV equivalence and extensibility"],
]
story += [table_caption("1.1", "Primary project artifacts"), make_table(artifact_rows, [55*mm, 115*mm]), PageBreak()]
section(story, "1.3", "Architecture at a glance")
story += [architecture_diagram(), figure_caption("1.1", "End-to-end project architecture from sourced routes to app and optimizer outputs")]
story.append(P("The Excel workbook is the single engineer-editable input source. The loader resolves stable IDs, the validator checks schema and physical ranges, and the input-preparation function resamples the selected route to a 1 s grid. Both the MATLAB kernel and the SLX model consume the same resolved parameter set.", "BodyTextBook"))

# Chapter 2
chapter(story, "2", "Selectable Hybrid and Battery-Electric Architectures")
section(story, "2.1", "Energy topology")
story.append(P("The architecture contains two independent battery packs, two rear hub motors, a diesel engine-generator set, an auxiliary electrical load, and a resistor load bank. The topology has two deliberately separated electrical paths. Only the active battery connects to the traction DC bus; the genset connects only to the standby battery through a dedicated charger. The genset therefore cannot propel the vehicle or reduce active-battery traction demand. During braking, regenerated electrical power supplies auxiliaries first, charges only the active battery second, and is sent to the resistor load bank if the active battery cannot accept the remainder.", "BodyTextBook"))
story += [energy_flow_diagram(), figure_caption("2.1", "Conceptual energy flow, isolated genset path, and regeneration priorities")]
section(story, "2.2", "Simulink decomposition")
simulink_image = image_flow(ASSET / "simulink_top_level.png", max_width=170*mm, max_height=72*mm)
story += [simulink_image, figure_caption("2.2", "Top-level Simulink model and logged signal fan-out")]
subsystems = [
    ["Subsystem", "Responsibility"],
    ["Route_and_Environment", "Publishes route speed, grade, and auxiliary multiplier"],
    ["Vehicle_Longitudinal_Dynamics", "Computes filtered acceleration and required wheel power"],
    ["Rear_Hub_Motor_Drive", "Applies motor torque, power, speed, efficiency, and regeneration limits"],
    ["Auxiliary_Loads", "Calculates base and HVAC electrical demand"],
    ["Supervisory_Energy_Management", "Enforces the 30% role swap, routes fixed genset power only to standby, and applies auxiliary-active-dump regeneration priority"],
    ["Battery_Pack_1 / Battery_Pack_2", "Integrate independent energy states under power and SOE limits"],
    ["Engine_Genset", "Produces constant optimum electrical power when enabled and calculates fuel rate"],
    ["Energy_and_Cost_Accounting", "Includes resistor-bank dissipation in the DC balance and accumulates energy/cost quantities"],
    ["Output_Logging", "Publishes stable workspace signals for inspection"],
]
story += [table_caption("2.1", "Top-level model responsibilities"), make_table(subsystems, [62*mm, 108*mm])]
section(story, "2.3", "Reference kernel versus SLX")
story.append(P("The ordinary-block SLX is intended for architecture inspection, signal tracing, and interactive simulation. The MATLAB kernel contains the complete deterministic limit handling and supervisory behavior used by batch studies and the optimizer. This separation keeps the teaching model editable while making the numerical reference path easy to test.", "BodyTextBook"))
section(story, "2.4", "Battery-electric alternative")
story.append(P("The Powertrain Architecture switch selects Hybrid or BEV without changing the component database. In BEV mode the diesel fuel, engine, generator, and isolated standby charger are absent. The former Use both batteries checkbox has been removed. A Battery set multiplier now defines installed capacity: 0.5, 1.0, 1.5, and 2.0 BEV sets mean one, two, three, and four connected packs. Odd pack totals assign the additional pack to the Battery 1 selection. All connected packs begin at the same SOE, 85% by default.", "BodyTextBook"))
story += [image_flow(ROOT / "results" / "BEV_App_Dual_Battery_View.png", max_width=170*mm, max_height=102*mm), figure_caption("2.3", "Selectable BEV architecture with scalable parallel battery banks, external charging, and regenerative return")]
story.append(P("For a requested signed DC power Pdc, the controller computes each connected bank's instantaneous charge or discharge capability from its SOE window, scaled pack count, power rating, derating factor, efficiency, and remaining energy headroom. Accepted power is shared in proportion to those capabilities:", "BodyTextBook"))
story += [equation("P<sub>1</sub> = P<sub>accepted</sub> C<sub>1</sub>/(C<sub>1</sub>+C<sub>2</sub>), &nbsp; P<sub>2</sub> = P<sub>accepted</sub> C<sub>2</sub>/(C<sub>1</sub>+C<sub>2</sub>)", "2.1")]
story.append(P("Regeneration retains the same first-principles destination order: auxiliaries first, connected battery pack or packs second, and the resistor load bank third. The BEV kernel sets genset power, genset starts, fuel rate, and fuel use identically to zero. HybridBus_BEVModel.slx is a separate editable fixed-step representation using the same database-resolved route, motor, driveline, auxiliary, battery, vehicle, and economic parameters.", "BodyTextBook"))
story += [image_flow(ROOT / "results" / "HybridBus_BEVModel_TopLevel.png", max_width=170*mm, max_height=68*mm), figure_caption("2.4", "Separate HybridBus_BEVModel.slx with parallel battery energy management and explicit zero fuel/genset path")]
section(story, "2.5", "Battery-set interpretation in Hybrid mode")
story.append(P("Hybrid accepts only positive whole-number set multipliers. One set contains one Battery 1 pack in one role bank and one Battery 2 pack in the alternate role bank. A multiplier of two therefore produces two Battery 1 packs in parallel and two Battery 2 packs in parallel. Exactly one bank is active for traction while the equal-sized alternate bank is standby; the 30% role-swap rule transfers those bank roles, not individual packs.", "BodyTextBook"))
story += [
    equation("N<sub>B1,Hybrid</sub> = N<sub>B2,Hybrid</sub> = k, &nbsp; k = 1,2,3,...", "2.2"),
    equation("N<sub>packs,BEV</sub> = 2k, &nbsp; k = 0.5,1.0,1.5,...", "2.3"),
]

# Chapter 3
chapter(story, "3", "Data Architecture and Reproducible Inputs")
section(story, "3.1", "Workbook plus modular engineering data")
story.append(P(f"The version {dashboard.get('DatabaseVersion')} workbook retains dashboard selections, common vehicle and control calibration, prices, tyre and final-drive data, and governance sheets. Route histories are stored one route per MAT file. Each battery and motor variant is one MATLAB data function, while each genset MATLAB file contains the matched genset, engine, generator, fuel map, and generator-efficiency map. IDs rather than row numbers form every interface.", "BodyTextBook"))
sheet_rows = [["Data group", "Storage", "Purpose"],
              ["Selections and calibration", "Workbook Dashboard, Vehicle, Environment, Control", "Chosen IDs, SOEs, prices, vehicle constants, and controls"],
              ["Routes", "data/routes/*.mat", "Independent time, distance, grade, geometry, elevation, provenance, and license records"],
              ["Battery variants", "data/batteries/*.m", "SOE/temperature current, OCV, resistance, energy, mass, and compatibility data"],
              ["Motor variants", "data/motors/*.m", "Torque-speed limits, two-dimensional loss maps, mass, and compatibility data"],
              ["Genset assemblies", "data/gensets/*.m", "Matched engine, generator, fuel, efficiency, and assembly data"],
              ["Common hardware", "Workbook Tyre and Final Drive", "Shared driveline ratings and efficiency"],
              ["Governance", "Workbook definitions/change log + extension guide", "Schema, units, provenance, and safe add-on procedure"]]
story += [table_caption("3.1", "Hybrid project data organization"), make_table(sheet_rows, [39*mm, 57*mm, 74*mm], font_size=6.8)]
section(story, "3.2", "Input pipeline")
story += [pipeline_diagram(), figure_caption("3.1", "Route and component data pipeline")]
story.append(P("The loader imports workbook calibration, recursively discovers component MATLAB files, and reads every route MAT file. Validation checks schemas, case-insensitive unique IDs, cross-references, lookup dimensions and monotonic breakpoints, physical ranges, route finiteness, monotonic time and distance, and nonzero mission length. Preparation resolves the selected records and resamples speed, grade, elevation, and auxiliary multiplier at the model sample time.", "BodyTextBook"))
section(story, "3.3", "Default configuration")
default_rows = [
    ["Input", "Selection / value", "Unit or interpretation"],
    ["Route", str(dashboard.get("SelectedRoute")), "stable route ID"],
    ["Battery 1 / 2", f"{dashboard.get('SelectedBattery1')} / {dashboard.get('SelectedBattery2')}", "independent packs"],
    ["Motor / genset", f"{dashboard.get('SelectedMotor')} / {dashboard.get('SelectedGenset')}", "two identical hub motors + one genset"],
    ["Battery set multiplier / load", f"{float(dashboard.get('BatterySetMultiplier')):.1f} / {float(dashboard.get('LoadMass_t')):.1f}", "sets / tonnes"],
    ["Calculated curb / total mass", f"{summary['CalculatedCurbMass_kg']/1000:.3f} / {summary['EstimatedVehicleMass_kg']/1000:.3f}", "tonnes"],
    ["Initial B1 / B2 SOE", f"{100*float(dashboard.get('InitialBattery1SOE')):.0f} / {100*float(dashboard.get('InitialBattery2SOE')):.0f}", "%"],
    ["Fuel / electricity price", f"{float(dashboard.get('FuelPrice')):.2f} / {float(dashboard.get('ElectricityPrice')):.2f}", "EUR/L / EUR/kWh"],
]
story += [table_caption("3.2", "Default case resolved from the Dashboard"), make_table(default_rows, [47*mm, 63*mm, 60*mm])]

# Chapter 4
chapter(story, "4", "Longitudinal Vehicle Dynamics")
section(story, "4.1", "Speed, acceleration, and grade")
story.append(P("The route speed is prescribed. A first-order discrete filter is applied to the raw finite-difference acceleration to reduce numerical spikes without introducing a continuous state.", "BodyTextBook"))
story += [
    equation("a<sub>raw,k</sub> = (v<sub>k</sub> - v<sub>k-1</sub>) / Δt<sub>k-1</sub>", "4.1"),
    equation("a<sub>k</sub> = a<sub>k-1</sub> + [Δt / (τ + Δt)] (a<sub>raw,k</sub> - a<sub>k-1</sub>)", "4.2"),
]
story.append(P("Road grade is stored as percent. The model converts it to an angle using θ = atan(grade/100). Headwind is added to vehicle speed and clipped so aerodynamic relative speed cannot become negative.", "BodyTextBook"))
section(story, "4.2", "Road-load forces")
story += [
    equation("F<sub>inertia</sub> = m a", "4.3"),
    equation("F<sub>roll</sub> = m g C<sub>rr</sub> cos(θ)", "4.4"),
    equation("F<sub>grade</sub> = m g sin(θ)", "4.5"),
    equation("F<sub>aero</sub> = 0.5 ρ C<sub>d</sub> A v<sub>air</sub><super>2</super>", "4.6"),
    equation("F<sub>tractive</sub> = F<sub>inertia</sub> + F<sub>roll</sub> + F<sub>grade</sub> + F<sub>aero</sub>", "4.7"),
    equation("P<sub>wheel,demand</sub> = F<sub>tractive</sub> v / 1000", "4.8"),
]
story.append(callout("Mass definition", "The fixed mass selector has been retired. Calculated curb mass equals a 15,000 kg base vehicle plus every installed battery pack and, in Hybrid mode only, the selected complete genset assembly. Total vehicle mass equals calculated curb mass plus the user-entered passenger, luggage, and cargo load. Motors and driveline are treated as part of the 15-tonne base and are not added again.", "info"))
story += [
    equation("m<sub>curb</sub> = 15000 + N<sub>B1</sub>m<sub>B1</sub> + N<sub>B2</sub>m<sub>B2</sub> + I<sub>Hybrid</sub>m<sub>genset</sub>", "4.9"),
    equation("m<sub>total</sub> = m<sub>curb</sub> + 1000 L<sub>tonnes</sub>", "4.10"),
]
section(story, "4.3", "Force scaling intuition")
story.append(P("Inertia, rolling resistance, and grade forces scale linearly with total mass. Aerodynamic force does not. Therefore mass strongly changes stop-and-go and climbing demand, while high-speed cruise also depends heavily on frontal area and drag coefficient.", "BodyTextBook"))
mass_fig = line_chart(
    mass_sweep["TotalVehicleMass_kg"] / 1000,
    [("Grid energy", mass_sweep["GridEnergy_kWh"], BLUE), ("Unmet energy", mass_sweep["UnmetEnergy_kWh"], RED)],
    "Mass sweep on the VECTO Urban route", "Total vehicle mass (tonnes)", "Energy (kWh)", y_min=0,
)
story += [mass_fig, figure_caption("4.1", "Calculated-mass sensitivity from 19 to 60 tonnes by varying entered load")]
story.append(P("The historical 19-to-60 tonne catalog is retained only as a reproducible sweep grid. Each study point is now realized by calculating the load required above the selected hardware-dependent curb mass. The default component set remains feasible at the low end, but unmet traction energy grows with total mass. Cost per kilometre alone must not rank infeasible cases because unsupplied energy is not purchased.", "BodyTextBook"))
section(story, "4.4", "Achieved-speed and performance formulations")
story.append(P("Backward mode treats route speed as achieved and computes required force. Constrained mode instead limits positive acceleration to the force that remains after road load, motor torque-speed/power, driveline, battery-current, terminal-voltage, and available-energy constraints. Terrain and auxiliary demand are sampled at actual travelled distance, speed is prevented from becoming negative, and a depleted or force-deficient vehicle remains stopped. Performance mode uses a proportional driver, explicit component limits, and a forward longitudinal plant to calculate achieved speed and distance.", "BodyTextBook"))
story += [
    equation("F<sub>acc,available</sub> = max[0, min(F<sub>motor</sub>, F<sub>battery</sub>) - F<sub>road</sub>]", "4.11"),
    equation("a<sub>achieved</sub> = min[a<sub>requested</sub>, F<sub>acc,available</sub>/m]", "4.12"),
    equation("v<sub>k+1</sub> = max[0, v<sub>k</sub> + a<sub>achieved,k</sub> Δt]", "4.13"),
    equation("s<sub>k+1</sub> = s<sub>k</sub> + 0.5(v<sub>k</sub>+v<sub>k+1</sub>)Δt", "4.14"),
]
story.append(callout("Performance evidence", "The KPI Vehicle Performance sub-tab reports desired versus achieved speed, route completion, tracking compliance, speed adequacy, RMS/maximum speed error, time below target, limiting cause, and termination reason. Repeat-route studies remain range/depletion experiments rather than single-pass tracking tests.", "info"))

# Chapter 5
chapter(story, "5", "Rear Hub Motors and Fixed Driveline")
section(story, "5.1", "Kinematics")
story += [
    equation("ω<sub>wheel</sub> = v / r<sub>loaded</sub>", "5.1"),
    equation("ω<sub>motor</sub> = i<sub>fd</sub> ω<sub>wheel</sub>,    n<sub>motor</sub> = 60 ω<sub>motor</sub> / (2π)", "5.2"),
]
story.append(P("The model uses two identical rear hub motors. The selected fixed reduction ratio connects wheel speed to motor speed. If motor speed exceeds the catalog maximum, available motor power is set to zero and the shortfall is logged.", "BodyTextBook"))
section(story, "5.2", "Torque and power envelope")
story += [
    equation("T<sub>available</sub> = min[T<sub>peak</sub>, P<sub>peak</sub> / max(ω<sub>motor</sub>, ε)]", "5.3"),
    equation("P<sub>tractive,max</sub> = 2 min(P<sub>peak</sub>, T<sub>available</sub> ω<sub>motor</sub>) η<sub>fd,mot</sub>", "5.4"),
]
selected_motor = motors[motors["ComponentID"] == str(dashboard.get("SelectedMotor"))].iloc[0]
speeds = list(range(0, int(selected_motor["MaxSpeed_rpm"]) + 1, 100))
torques = []
for rpm in speeds:
    omega = rpm * 2 * math.pi / 60
    t = float(selected_motor["PeakTorque_Nm"])
    if rpm > float(selected_motor["BaseSpeed_rpm"]):
        t = min(t, float(selected_motor["PeakPower_kW"]) * 1000 / max(omega, 1))
    torques.append(t)
motor_fig = line_chart(speeds, [("Peak envelope", torques, BLUE)], f"{selected_motor['ComponentID']} torque-speed envelope", "Motor speed (rpm)", "Torque per motor (Nm)", y_min=0)
story += [motor_fig, figure_caption("5.1", "Selected motor constant-torque and constant-power regions")]
section(story, "5.3", "Directional efficiency and regeneration")
story.append(P("Traction wheel power is divided by motor and final-drive motoring efficiencies to obtain DC demand. During braking, delivered negative wheel power is multiplied by regeneration efficiencies, producing negative motor DC power. Regeneration is limited by motor capability and fixed-drive direction. The Hybrid limit uses the currently active scaled bank capability; BEV sums the instantaneous capability of every connected parallel pack represented by the set multiplier.", "BodyTextBook"))
story += [
    equation("P<sub>motor,dc</sub> = P<sub>wheel</sub> / (η<sub>fd,mot</sub> η<sub>motor,mot</sub>)    for traction", "5.5"),
    equation("P<sub>motor,dc</sub> = P<sub>wheel</sub> η<sub>fd,reg</sub> η<sub>motor,reg</sub>    for regeneration", "5.6"),
]

# Chapter 6
chapter(story, "6", "Battery Energy Model")
section(story, "6.1", "Energy state and SOE")
story.append(P("Each battery retains usable stored energy as its dynamic state and uses a first-order Thevenin capability layer. Open-circuit voltage, internal resistance, maximum discharge current, and maximum charge current are two-dimensional functions of SOE and ambient battery temperature. This supports state-dependent terminal voltage, current, power capability, and ohmic loss without claiming a full electrothermal cell model.", "BodyTextBook"))
story += [
    equation("SOE = E / E<sub>usable</sub>", "6.1"),
    equation("E<sub>k+1</sub> = E<sub>k</sub> - P<sub>dis</sub> Δt / (3600 η<sub>dis</sub>)", "6.2"),
    equation("E<sub>k+1</sub> = E<sub>k</sub> + |P<sub>chg</sub>| η<sub>chg</sub> Δt / 3600", "6.3"),
]
section(story, "6.2", "Current, voltage, and energy limits")
story.append(P("At every time step the battery helper interpolates the SOE-temperature current maps, OCV map, and resistance map. Map current is further limited by terminal-voltage headroom. Accepted power is then bounded by the resulting electrical capability and by the energy-implied power needed to remain within minimum or maximum SOE.", "BodyTextBook"))
story += [
    equation("I<sub>dis,max</sub> = min[I<sub>map,dis</sub>(SOE,T), (OCV(SOE,T)-V<sub>min</sub>)/R(SOE,T)]", "6.4"),
    equation("I<sub>chg,max</sub> = min[I<sub>map,chg</sub>(SOE,T), (V<sub>max</sub>-OCV(SOE,T))/R(SOE,T)]", "6.5"),
    equation("P<sub>dis,max</sub> = I<sub>dis,max</sub>[OCV-I<sub>dis,max</sub>R]/1000", "6.6"),
    equation("P<sub>chg,max</sub> = I<sub>chg,max</sub>[OCV+I<sub>chg,max</sub>R]/1000", "6.7"),
]
story.append(callout("Numerical safeguard", "The energy state is clamped to [MinSOE, MaxSOE] after each update. Regenerative power rejected by the active battery is assigned to the resistor load bank; a genset-charger mismatch remains a separately logged rejected-charge term. Unserved positive demand is recorded as unmet DC power.", "note"))
section(story, "6.3", "Hybrid active/standby roles versus BEV parallel packs")
story.append(P("The Hybrid supervisor permits only one battery bank to discharge for traction. This is intentionally conservative: the active bank must support residual DC demand while the equal-sized alternate bank remains available for standby charging and deterministic role switching. Bank energy, mass, charge, discharge, and regeneration limits scale linearly with the whole-number set multiplier. BEV mode removes active/standby roles and connects the installed one-or-more packs in parallel; each selected bank remains independently bounded by SOE, energy headroom, power, derating, and efficiency limits.", "BodyTextBook"))
battery_rows = [["Catalog range", "Minimum", "Maximum"]]
for label, col, unit in [
    ("Usable energy", "UsableEnergy_kWh", "kWh"),
    ("Nominal voltage", "NominalVoltage_V", "V"),
    ("Mass", "Mass_kg", "kg"),
]:
    battery_rows.append([label, f"{batteries[col].min():.0f} {unit}", f"{batteries[col].max():.0f} {unit}"])
story += [table_caption("6.1", "Synthetic battery catalog coverage"), make_table(battery_rows, [70*mm, 50*mm, 50*mm], alignments=["LEFT", "CENTER", "CENTER"])]

# Chapter 7
chapter(story, "7", "Engine-Generator Set and Fuel Model")
section(story, "7.1", "Constant best-efficiency operating point")
story.append(P("The genset has one permitted electrical command while on: the catalog OptimumPower_kW value, bounded by the rated maximum. It is not load-following and is not connected to the traction bus. Its dedicated charger sends power only to the pack currently designated standby. Battery charge-power and upper-SOE limits determine accepted power; any transient charger mismatch is recorded separately from regenerative load-bank dissipation.", "BodyTextBook"))
section(story, "7.2", "Generator and engine maps")
story += [
    equation("P<sub>mech</sub> = P<sub>gen,dc</sub> / η<sub>gen</sub>(λ)", "7.1"),
    equation("ṁ<sub>fuel</sub> = max[ṁ<sub>idle</sub>, P<sub>mech</sub> BSFC(λ) / (1000 ρ<sub>fuel</sub> 3600)]", "7.2"),
]
map_fig = line_chart(
    fuel_map["NormalizedEngineLoad"],
    [("BSFC (g/kWh)", fuel_map["BSFC_g_kWh"], ORANGE),
     ("Gen efficiency x 400", generator_map["Efficiency"] * 400, BLUE)],
    "Normalized engine and generator maps", "Normalized load", "Scaled map value", y_min=250, y_max=400,
)
story += [map_fig, figure_caption("7.1", "Synthetic BSFC and generator-efficiency maps used for interpolation")]
story.append(P("The generator efficiency is clipped to a physically reasonable numerical range, and BSFC is linearly interpolated over normalized load. A fixed start-fuel penalty is added on every off-to-on transition.", "BodyTextBook"))
section(story, "7.3", "What the fuel model does not claim")
story.append(callout("Not an emissions model", "There is no engine warm-up, transient combustion, aftertreatment, ambient correction, fuel-temperature effect, or pollutant calculation. The output is a concept diesel-volume estimate, not an emissions or homologation result.", "warn"))

# Chapter 8
chapter(story, "8", "Supervisory Energy Management")
section(story, "8.1", "Control objectives")
story.append(P("In Hybrid mode the deterministic supervisor has five priorities: keep the genset electrically isolated from traction; preserve exactly one active traction pack; swap roles at 30% SOE when the alternate pack is ready; recharge the depleted standby pack at constant best-efficiency genset power; and allocate regeneration strictly to auxiliaries, then the active battery, then the resistor load bank. BEV mode uses the separate policy in Section 8.6.", "BodyTextBook"))
story += [supervisor_diagram(), figure_caption("8.1", "Supervisory operating modes and principal transitions")]
story.append(PageBreak())
section(story, "8.2", "Decision order")
story.append(numbered_list([
    "Read both pack SOEs and retain the current active-battery state.",
    "If active SOE is at or below 30% and the standby pack is above 30%, exchange the active and standby roles immediately.",
    "Start the genset when standby SOE is at or below 30%; while on, command the constant OptimumPower_kW value.",
    "Send positive motor-plus-auxiliary demand only to the active battery; the genset is excluded from this equation.",
    "Send genset power only to the standby battery; the genset never participates in the traction-node balance.",
    "For negative motor DC power, supply auxiliary demand first, send the remaining regenerative power only to the active battery, and divert any unaccepted remainder to the resistor load bank.",
    "Update energy states, fuel, operating mode, losses, load-bank heat, rejected genset charge, and the balance residual.",
]))
section(story, "8.3", "Regeneration allocation by first principles")
story.append(P("Let regenerative power be a non-negative available quantity. Auxiliary loads remain powered during braking, so they consume the first portion directly at the traction DC bus. Only the remainder is offered to the active pack. The battery helper then applies its instantaneous charge-power, regeneration-power, efficiency, and upper-SOE limits. The resistor load bank receives exactly the terminal power that the active battery cannot accept. The standby pack is excluded from this path because its only charging source is the isolated genset charger.", "BodyTextBook"))
story += [
    equation("P<sub>reg</sub> = max(0, -P<sub>motor,dc</sub>)", "8.1"),
    equation("P<sub>reg,aux</sub> = min(P<sub>reg</sub>, P<sub>aux</sub>)", "8.2"),
    equation("P<sub>reg,rem</sub> = max(0, P<sub>reg</sub> - P<sub>reg,aux</sub>)", "8.3"),
    equation("P<sub>dump</sub> = max(0, P<sub>reg,rem</sub> - P<sub>reg,active</sub>)", "8.4"),
    regeneration_priority_diagram(),
    figure_caption("8.2", "Strict auxiliary-active-battery-resistor regeneration hierarchy"),
]
regen_fig = line_chart(
    regen_priority["Time_s"],
    [("Available", regen_priority["Available_kW"], GREEN),
     ("To auxiliary", regen_priority["ToAuxiliary_kW"], GOLD),
     ("To active battery", regen_priority["ToActiveBattery_kW"], BLUE),
     ("Resistor bank", regen_priority["ResistorLoadBank_kW"], DUMP)],
    "Full-active-battery downhill regeneration case", "Time (s)", "Power (kW)", y_min=0,
)
story += [regen_fig, figure_caption("8.3", "When the active battery is full, surplus regenerative power becomes controlled waste heat")]
story.append(callout("Thermal interpretation", "The resistor-bank signal is an electrical power sink. A production design would require continuous and transient thermal ratings, cooling airflow, temperature sensing, contactor diagnostics, insulation monitoring, and a braking fallback if the bank is unavailable. Those thermal and safety dynamics are outside the present concept model.", "warn"))
section(story, "8.4", "Hysteresis and dwell time")
story.append(P("The 30% role threshold is an architectural rule rather than a tunable calibration. A role swap is allowed only if the alternate pack is above 30%; this prevents rapid swapping when both packs are depleted. If both are low, the current active pack remains connected while the genset charges standby. Charging continues until the standby pack reaches its upper SOE limit or becomes the active pack.", "BodyTextBook"))
mode_rows = [
    ["Mode", "Meaning", "Primary power behavior"],
    ["1", "Battery 1 active", "B1 alone supplies positive traction-bus demand"],
    ["2", "Battery 2 active", "B2 alone supplies positive traction-bus demand"],
    ["3", "B1 active; B2 charged", "Genset fixed output is routed only to standby B2"],
    ["4", "B2 active; B1 charged", "Genset fixed output is routed only to standby B1"],
    ["6", "Regeneration", "Auxiliary first, active battery second, resistor load bank third"],
    ["7", "Protection", "Sources are clamped and unmet demand is recorded"],
    ["8", "BEV single-pack traction", "A 0.5 set connects one Battery 1 pack; Battery 2 count is zero"],
    ["9", "BEV parallel traction", "All packs represented by the set multiplier share accepted DC power"],
    ["10", "BEV regeneration", "Auxiliary first, connected pack(s) second, resistor bank third"],
]
story += [table_caption("8.1", "Operating-mode interpretation"), make_table(mode_rows, [18*mm, 55*mm, 97*mm], alignments=["CENTER", "LEFT", "LEFT"])]
section(story, "8.5", "Repeated-route depletion study")
story.append(P("The Configuration panel includes a Repeat route until depleted checkbox. When cleared, the solver executes exactly one selected mission, preserving the standard comparison workflow. When selected, the time-speed-grade-auxiliary sequence is concatenated continuously while distance and time remain monotonic. This converts a route case into a range-to-depletion experiment without changing the component physics or supervisory priorities.", "BodyTextBook"))
story.append(P("The genset consumes no more than the configured fuel-tank volume. After fuel is exhausted, the vehicle continues from available battery energy. The repeated mission terminates only when the active pack reaches its minimum usable-energy bound, the alternate pack is not above the 30% role-swap threshold, and positive DC demand is unmet. Thus the stop condition means that no admissible traction-energy source remains; it is not an arbitrary route-count limit.", "BodyTextBook"))
story += [
    equation("V<sub>fuel,used</sub> &ge; V<sub>tank</sub>", "8.5"),
    equation("E<sub>active</sub> &le; E<sub>active,min</sub> &nbsp; and &nbsp; SOE<sub>alternate</sub> &le; 30%", "8.6"),
]
story.append(callout("Interpretation", "A repeated-route result is a controlled range estimate for the prescribed repeating mission. It is not a real timetable, charging plan, driver-hours model, thermal endurance result, or warranty prediction.", "warn"))
section(story, "8.6", "BEV supervisory mode")
story.append(P("BEV control bypasses the 30% Hybrid role-swap and every genset state. The controller reads deterministic pack counts from the half-step battery-set multiplier, forms total traction-node demand, computes instantaneous scaled-bank capabilities, and distributes accepted discharge or charge power. If requested positive DC power exceeds connected capability, the difference is logged as unmet DC power. During braking, the same auxiliary-battery-resistor priority is applied across every connected pack.", "BodyTextBook"))
story += [
    equation("P<sub>BEV,accepted</sub> = min(|P<sub>request</sub>|, Σ<sub>j connected</sub>C<sub>j</sub>)", "8.7"),
    equation("N<sub>B1</sub> = ceil(N<sub>packs</sub>/2), &nbsp; N<sub>B2</sub> = floor(N<sub>packs</sub>/2)", "8.8"),
]
story.append(callout("BEV invariants", "All connected packs start at the same SOE (85% by default). With 0.5 set, Battery 2 count and power are exactly zero. For every BEV multiplier, genset power, fuel rate, fuel consumption, runtime, and starts are exactly zero.", "note"))
section(story, "8.7", "Ordered BEV-then-Hybrid comparison")
story.append(P("A default-off Configuration checkbox provides a controlled two-run experiment. When cleared, Run Manual executes only the architecture selected by the Hybrid-BEV slider. When checked, the orchestrator executes BEV first and Hybrid second with the same route, components, load, prices, auxiliaries, and repeat-route setting. The BEV case makes connected starting SOEs equal to the displayed Battery 1 value; the Hybrid case retains its configured B1/B2 values. Both results remain in memory, Hybrid histories are displayed last, and the status and Simulation Analysis header report both operating-cost outcomes.", "BodyTextBook"))
story.append(callout("Comparable battery-set constraint", "The paired run requires a positive whole-number battery-set multiplier because that domain is valid for both architectures. A fractional BEV set is never silently rounded for the Hybrid case; the app reports a clear configuration error instead.", "warn"))

# Chapter 9
chapter(story, "9", "Power, Energy, Cost, and Range Accounting")
section(story, "9.1", "DC power conservation")
story.append(P("Two node equations make the isolation explicit. At the traction node, only the active battery balances motor power, auxiliary demand, and resistor-bank dissipation. At the charger node, the genset balances standby-battery charging and any transient rejected charger power. Summing the nodes gives the project-wide residual without implying a physical genset-to-traction connection.", "BodyTextBook"))
story += [
    equation("P<sub>active</sub> - P<sub>motor,dc</sub> - P<sub>aux</sub> + P<sub>unmet</sub> - P<sub>dump</sub> = 0", "9.1a"),
    equation("P<sub>gen</sub> + P<sub>standby</sub> - P<sub>gen,rejected</sub> = 0", "9.1b"),
    equation("P<sub>gen</sub> + P<sub>B1</sub> + P<sub>B2</sub> - P<sub>motor,dc</sub> - P<sub>aux</sub> + P<sub>unmet</sub> - P<sub>dump</sub> - P<sub>gen,rejected</sub> = P<sub>residual</sub>", "9.1c"),
    equation("P<sub>rejected</sub> = P<sub>dump</sub> + P<sub>gen,rejected</sub>", "9.1d"),
    equation("Σ<sub>j connected</sub>P<sub>Bj</sub> - P<sub>motor,dc</sub> - P<sub>aux</sub> + P<sub>unmet</sub> - P<sub>dump</sub> = P<sub>residual</sub> &nbsp; (BEV)", "9.1e"),
]
section(story, "9.2", "Equivalent grid replenishment")
story += [
    equation("E<sub>grid</sub> = Σ<sub>connected packs</sub> max(0,E<sub>i,0</sub>-E<sub>i,f</sub>) / η<sub>grid</sub>", "9.2"),
    equation("C<sub>total</sub> = V<sub>fuel</sub> c<sub>fuel</sub> + E<sub>grid</sub> c<sub>electricity</sub>", "9.3"),
    equation("C<sub>km</sub> = C<sub>total</sub> / d<sub>route</sub>", "9.4"),
]
story.append(P("Terminal battery surplus is not credited by default. This prevents an artificial negative electric cost and makes the equivalent-replenishment comparison conservative.", "BodyTextBook"))
section(story, "9.3", "Range estimates")
story += [
    equation("R<sub>battery</sub> = E<sub>usable,initial</sub> / e<sub>DC,route</sub>", "9.5"),
    equation("R<sub>fuel</sub> = V<sub>tank</sub> / v<sub>fuel,per-km</sub>", "9.6"),
]
story.append(callout("Range is conditional", "Range estimates assume that the observed route energy intensity and supervisory behavior remain representative. If the example route never starts the genset, the code estimates supported fuel range at the selected optimum genset point rather than reporting infinity.", "warn"))
section(story, "9.4", "Feasibility before economics")
story.append(P("A low cost per kilometre is meaningful only when unmet traction energy is negligible and the balance residual satisfies tolerance. An infeasible case can look artificially cheap because energy that the vehicle failed to deliver was never purchased. The optimizer therefore filters feasibility before ranking cost.", "Lead"))

# Chapter 10
chapter(story, "10", "Numerical Implementation in MATLAB and Simulink")
section(story, "10.1", "Discrete solution")
story.append(P("The default solver interpretation is fixed-step discrete with a 1 s sample time. Route speed, grade, and auxiliary multiplier are linearly interpolated; stop flags use previous-value interpolation. Energy updates use time-step integration in kWh, while fuel rate is integrated in litres.", "BodyTextBook"))
section(story, "10.2", "Execution sequence")
run_code = """
cd('C:\\TempData\\Hybrid_Vehicle\\HybridBusProject')
project = startup_hybrid_bus;
db = load_hybrid_bus_database(fullfile('data','HybridBus_ComponentDatabase.xlsx'));
Results = run_hybrid_bus_simulation;
Sequence = run_powertrain_sequence(db, struct('PowertrainMode',"Hybrid"), true);
open_system('HybridBus_BackwardModel.slx')
open_system('HybridBus_BEVModel.slx')
app = HybridBusApp;
"""
story.append(code_block(run_code))
story.append(P("The runner loads and validates the database, resolves the selected input structure, executes the detailed kernel, attaches validation metadata, and optionally exports MAT and CSV files. The SLX workspace publisher assigns only the stable, unit-suffixed variables needed by the block diagram.", "BodyTextBook"))
section(story, "10.3", "Results structure")
result_rows = [
    ["Field", "Contents"],
    ["Metadata", "database file/version, MATLAB release, model version, timestamp"],
    ["SelectedConfiguration", "stable IDs for route and every selected component/calibration"],
    ["InputParameters", "fully resolved scalar structures, maps, and resampled route"],
    ["Signals", "grouped vehicle, wheel, motor, auxiliary, regeneration, battery, genset, controller, and energy histories"],
    ["Summary", "distance, fuel, grid energy, regeneration allocation, load-bank heat, costs, range, SOE, unmet energy, residual, and mass decomposition"],
    ["Validation", "warnings, constraint violations, and feasibility flag"],
]
story += [table_caption("10.1", "Simulation output contract"), make_table(result_rows, [50*mm, 120*mm])]

# Chapter 11
chapter(story, "11", "Using the Explorer App")
section(story, "11.1", "Configuration panel")
story.append(P("The programmatic UIFigure app groups route, editable load, calculated curb and total mass, and auxiliary selection in Mission Inputs. Battery, motor, and genset selections remain directly below. Initial battery SOEs are entered as percentages and converted to fractions only at the model boundary. Fuel and electricity prices show EUR/L and EUR/kWh. The battery-set multiplier controls installed bank count; Repeat route until depleted selects a range experiment; and Run BEV first, then Hybrid selects the ordered architecture comparison defined in Section 8.7.", "BodyTextBook"))
story.append(P("The Powertrain Architecture tab opens first and includes a blue/gray Hybrid-BEV switch. In Hybrid it shows the isolated standby-charging chain, active-battery selector, traction DC bus, regenerative return, auxiliary-first branch, and resistor load bank. In BEV it replaces the fuel and genset chain with external charging and scalable parallel battery banks. Every component block is clickable: a non-modal window reports selected catalog IDs, ratings, units, implemented role, control rules, mass contribution, and concept-model limitations.", "BodyTextBook"))
app_image = image_flow(ASSET / "hybrid_bus_app.png", max_width=170*mm, max_height=102*mm)
story += [app_image, figure_caption("11.1", "Current Explorer Configuration panel with calculated mass; the default-off BEV-then-Hybrid control is selected here for illustration")]
architecture_app_image = image_flow(ASSET / "hybrid_bus_app_architecture.png", max_width=170*mm, max_height=102*mm)
story += [architecture_app_image, figure_caption("11.2", "Architecture tab with the wheel-to-inverter regenerative return and numbered destination priorities")]
story += [image_flow(ROOT / "documentation" / "screenshots" / "architecture_clickable_blocks.png", max_width=170*mm, max_height=102*mm), figure_caption("11.3", "Clickable architecture blocks expose the underlying selected component specifications")]
route_map_image = image_flow(ASSET / "app_mannheim_city_map.png", max_width=170*mm, max_height=102*mm)
story += [route_map_image, figure_caption("11.4", "Route Map 2D view displaying stored latitude/longitude geometry and endpoint markers")]
story.append(P("The Route Map switch preserves the geographic 2D view and reveals a second Elevation-Slope switch only in 3D mode. Elevation is plotted in metres from the stored route geometry. Slope is derived from successive elevation change divided by geographic path-distance change and displayed in percent. Blue identifies the selected side of every slider-style switch; gray identifies the alternative.", "BodyTextBook"))
route_3d_views = Table([[image_flow(ROOT / "documentation" / "screenshots" / "route_map_3d_elevation_toggle.png", max_width=82*mm, max_height=58*mm), image_flow(ROOT / "documentation" / "screenshots" / "route_map_3d_slope_toggle.png", max_width=82*mm, max_height=58*mm)]], colWidths=[85*mm,85*mm])
route_3d_views.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0)]))
story += [route_3d_views, figure_caption("11.5", "Route Map 3D quantity switch: elevation in metres (left) and road slope in percent (right)")]
section(story, "11.2", "Manual-study and architecture-comparison workflow")
story.append(numbered_list([
    "Select the database and confirm that the status reports a validated version.",
    "Choose the route, load in tonnes, auxiliary profile, components, battery-set multiplier, initial SOEs, and prices; confirm the calculated curb and total masses.",
    "Leave Repeat route until depleted clear for a one-cycle comparison, or select it deliberately for a range-to-depletion study.",
    "Leave Run BEV first, then Hybrid clear to execute the architecture selected by the slider, or check it to execute the ordered paired comparison.",
    "Run the case and inspect speed, power flow, battery SOE, genset/fuel, KPIs, Simulation Analysis, and warnings. After a paired run, the plots show Hybrid while the header preserves both cost outcomes.",
    "Treat any unmet traction energy as a failed design case, not as an acceptable low-cost solution.",
    "Export the selected result only after checking assumptions and feasibility.",
]))
comparison_rows = [["Mode", "Mass (t)", "Cost (EUR/km)", "Fuel (L)", "Grid eq. (kWh)", "Feasible"]]
for _, row in powertrain_comparison.iterrows():
    comparison_rows.append([
        row["Mode"], f"{row['VehicleMass_kg']/1000:.3f}", f"{row['CostPer_km']:.4f}",
        f"{row['Fuel_L']:.2f}", f"{row['GridEquivalentEnergy_kWh']:.2f}",
        "Yes" if bool(row["Feasible"]) else "No",
    ])
story += [table_caption("11.1", "Default ordered BEV-then-Hybrid manual comparison"), make_table(comparison_rows, [24*mm, 26*mm, 34*mm, 25*mm, 38*mm, 23*mm], font_size=6.8, alignments=["LEFT","RIGHT","RIGHT","RIGHT","RIGHT","CENTER"])]
story.append(callout("Reading the paired result", "This is a controlled architecture comparison, not an optimization. Component IDs and mission inputs are held fixed, but architecture-specific mass and starting-SOE rules still apply. The BEV excludes genset mass and fuel; Hybrid includes the selected genset and active/standby starting states.", "note"))
section(story, "11.3", "Management KPI dashboard")
story.append(P("The KPIs tab is organized into Executive Decision, Engineering Scorecard, Vehicle Performance, and Robustness and Range sub-tabs. Executive Decision compares BEV and Hybrid outcomes for the selected route, load, auxiliaries, prices, and run scope and provides a qualified recommendation. Engineering Scorecard exposes completion, energy, power, SOE, conservation, and cost gates. Vehicle Performance distinguishes prescribed demand from achieved-speed evidence. Robustness and Range compares operating cost, source energy, range/depletion behavior, and decision sensitivity. Single-mode, ordered BEV-then-Hybrid, and repeat-until-depleted runs are interpreted without inventing unavailable comparison evidence.", "BodyTextBook"))
kpi_dashboard_image = image_flow(ASSET / "kpi_dashboard.png", max_width=170*mm, max_height=102*mm)
story += [kpi_dashboard_image, figure_caption("11.6", "Management KPI dashboard with icon-led values, units, configuration context, and feasibility status")]
section(story, "11.4", "Simulation Analysis tab")
story.append(P("Simulation Analysis is an interpretation layer, not another solver. Its energy-allocation chart compares grid-equivalent energy, genset electrical output, recovered regeneration, auxiliaries, battery throughput, and load-bank waste. The duty-share chart reports the fraction of simulated time for which each battery held the active traction role. The assessment table converts principal limits and balances into concise engineering statements covering mission completion, SOE windows, regeneration utilization, dump energy, genset runtime and starts, unmet demand, and conservation error.", "BodyTextBook"))
analysis_image = image_flow(ASSET / "simulation_analysis.png", max_width=170*mm, max_height=102*mm)
story += [analysis_image, figure_caption("11.7", "Simulation Analysis tab combining energy allocation, battery duty share, and engineering assessment")]
story.append(callout("Management versus engineering views", "Use the KPI dashboard to communicate outcomes quickly. Use Simulation Analysis to explain why those outcomes occurred and whether a limit, energy sink, or control transition requires attention.", "note"))
section(story, "11.5", "Model Credibility tab")
story.append(P("The Model Credibility tab prevents evidence from being compressed into one misleading green status. Concept verification, implementation equivalence, supplier calibration, and measured-vehicle validation are separate gates. Requirements traceability, the behavioral suite, deterministic sensitivity screening, and both Hybrid and BEV MATLAB-Simulink equivalence currently pass, while supplier calibration and measured-vehicle validation remain unavailable. The tab keeps every PASS, FAIL, NOT AVAILABLE, and NOT IMPLEMENTED state visible with its evidence and required decision.", "BodyTextBook"))
credibility_image = image_flow(ASSET / "model_credibility.png", max_width=170*mm, max_height=102*mm)
story += [credibility_image, figure_caption("11.8", "Model Credibility tab separating verified concept behavior from open implementation and validation gates")]
section(story, "11.6", "Signals and Detailed Plot axes")
story.append(P("Signals and Detailed Plot share a Time-Distance switch. Time remains the internal independent variable in seconds, but display units adapt to mission duration: minutes for missions shorter than two hours and hours for longer missions. Distance mode uses cumulative vehicle distance in kilometres, so users can relate power, SOE, torque, and speed events directly to route position. The Battery SOE axes are always displayed in percent.", "BodyTextBook"))
story += [image_flow(ROOT / "documentation" / "screenshots" / "signals_distance_axis.png", max_width=170*mm, max_height=102*mm), figure_caption("11.9", "Signals tab using cumulative distance in kilometres as the common x-axis")]
section(story, "11.7", "Optimization tab")
story.append(P("The optimization view ranks feasible configurations and keeps stable component IDs visible. Hybrid mode varies Battery 1, the motor pair, genset, and final drive while Battery 2 remains fixed. BEV mode varies both batteries, the motor pair, and final drive; the genset is excluded. Cancellation is checked between simulations, so the bounded search remains responsive without requiring Parallel Computing Toolbox. Chapter 12 defines the complete search and interpretation workflow.", "BodyTextBook"))

# Chapter 12
chapter(story, "12", "Configuration Optimization")
section(story, "12.1", "Purpose and study boundary")
story.append(P("The optimizer is a deterministic design-screening tool. It answers: among the component combinations actually evaluated for one prescribed mission, which feasible combination has the lowest modeled operating cost per kilometre? It does not redesign components, tune continuous parameters, predict hardware life, or prove a global optimum.", "BodyTextBook"))
fixed_varied_rows = [
    ["Mode / held fixed by the app", "Varied by the app"],
    ["Both: route, entered load, calculated mass, auxiliaries, environment, prices", "Both: motor pair and final drive"],
    ["Both: tyre, control calibration, fuel tank, repeat-route flag", "Hybrid: Battery 1 and genset"],
    ["Both: battery-set multiplier; Hybrid: Battery 2 and both initial SOEs", "BEV: Battery 1 and Battery 2"],
    ["BEV: equal initial SOE and connected pack count derived from multiplier", "BEV: no genset candidate"],
]
story += [table_caption("12.1", "Study variables controlled by the Optimize button"), make_table(fixed_varied_rows, [85*mm, 85*mm])]
story.append(callout("Mode-specific search", "Hybrid preserves the selected standby Battery 2 while varying Battery 1. BEV varies both pack selections because both may operate in parallel; the voltage-class compatibility rule is enforced before simulation.", "note"))

section(story, "12.2", "Candidate set and mixed-radix enumeration")
story.append(P("For every varied category, the optimizer reads stable component IDs from catalog rows marked OptimizationEnabled. If the four candidate counts are nB, nM, nG, and nF, the complete discrete search space is", "BodyTextBook"))
story += [equation("N<sub>space</sub> = n<sub>B</sub> n<sub>M</sub> n<sub>G</sub> n<sub>F</sub>", "12.1")]
story.append(P("A mixed-radix counter converts each linear evaluation index into one Battery 1, motor, genset, and final-drive combination. This avoids materializing a large Cartesian-product table in memory and gives repeatable ordering between runs with the same database.", "BodyTextBook"))
story.append(P("Before enumeration, any currently displayed component ID supplied by the app is moved to the front of its candidate list. The first candidate is therefore the complete displayed configuration, including the selected final drive. This release fixed a handoff defect in which the hidden final-drive ID was omitted: the 40-case budget could be consumed by an incompatible FD-01 prefix before the displayed valid FD-08 was reached. The complete configuration handoff now includes tyre, final drive, environment, control calibration, initial active battery, auxiliary scalar, and fuel-tank capacity.", "BodyTextBook"))
story += [optimization_diagram(), figure_caption("12.1", "Compatibility, simulation, feasibility, and ranking sequence")]

section(story, "12.3", "Meaning of Max configurations")
story.append(P("Max configurations is the evaluation budget for one click of Optimize. Its default value of 40 means that no more than forty candidate combinations are inspected in that run. Both compatible simulated cases and combinations rejected by the pre-simulation compatibility filter consume an evaluation number. The field is therefore a runtime-control setting, not a request for forty feasible solutions.", "BodyTextBook"))
story += [equation("N<sub>evaluated</sub> = min(N<sub>max</sub>, N<sub>space</sub>)", "12.2")]
story.append(P("Increasing the budget explores more of the deterministic catalog order and may find a lower-cost feasible design, but simulation time rises approximately with the number of compatible cases. If Nmax is smaller than Nspace, the reported winner is only the best among the evaluated prefix. A global discrete optimum is established only when the complete enabled search space is evaluated, or when a separately justified search strategy covers it.", "BodyTextBook"))
budget_rows = [
    ["Setting", "Interpretation", "Expected effect"],
    ["Small budget", "Rapid screening of the first enabled combinations", "Fast, but stronger risk of missing a better design"],
    ["40 (default)", "At most forty candidates are checked", "Practical interactive study size"],
    ["At least Nspace", "Every enabled combination is checked", "Exhaustive discrete ranking for the stated catalogs"],
]
story += [table_caption("12.2", "How to interpret the evaluation budget"), make_table(budget_rows, [34*mm, 69*mm, 67*mm])]

section(story, "12.4", "Compatibility filters")
story.append(P("Compatibility filtering removes combinations that violate static interface or rating constraints before the drive-cycle simulation. Rejected rows remain in the evaluated-results table with a reason, making the search auditable.", "BodyTextBook"))
story.append(bullet_list([
    "Battery-to-motor DC voltage class difference no greater than 50 V.",
    "Motor speed below maximum at the selected route speed, tyre radius, and final-drive ratio.",
    "Final-drive ratio inside the selected motor compatibility interval.",
    "Hybrid: at least one selected battery has adequate continuous discharge rating for the motor rating.",
    "BEV: summed connected-pack discharge capability covers the pair's continuous motor rating.",
    "Hybrid only: genset optimum power is within its rating and within either standby battery's charge-power capability.",
]))
story.append(PageBreak())
section(story, "12.5", "Simulation and feasibility gates")
story.append(P("Every compatible row is resolved into a complete model input and simulated over the selected route. A low cost is not sufficient. The row enters the feasible ranking only if it passes all three gates below.", "BodyTextBook"))
story += [equation("E<sub>unmet,traction</sub> &lt; 10<super>-3</super> kWh", "12.3")]
story += [equation("E<sub>balance,error</sub> ≤ E<sub>balance,tolerance</sub>", "12.4")]
story += [equation("|SOE<sub>combined,final</sub> - SOE<sub>combined,target</sub>| ≤ ΔSOE<sub>terminal</sub>", "12.5")]
story.append(P("The first gate rejects component sets that cannot satisfy requested wheel demand. The second detects inconsistent source, sink, or loss accounting. The third is applied to the Hybrid charge-sustaining comparison so a configuration cannot appear economical merely because it finishes with an excessive battery-energy deficit. A BEV mission is intentionally charge depleting, so the Hybrid terminal-SOE gate is inapplicable; BEV feasibility is governed by delivered mission energy, battery limits, and energy conservation. A rejection reason is stored when an applicable gate fails or a simulation raises an error.", "BodyTextBook"))

section(story, "12.6", "Objective, ranking, and tie-break")
story += [equation("minimize C<sub>km</sub> = (C<sub>fuel</sub> + C<sub>grid-equivalent electricity</sub>) / d<sub>route</sub>", "12.6")]
story.append(P("Only feasible rows are sorted. Operating cost per kilometre is the primary objective; fuel consumption in litres per 100 km is the ascending tie-break. The optimization table displays the highest-ranked feasible rows, up to ten. Stable catalog IDs are retained so a result can be reproduced from the workbook.", "BodyTextBook"))

section(story, "12.7", "What happens after the search")
story.append(numbered_list([
    "The status line reports the number of evaluated combinations; during the run it also shows the latest modeled cost.",
    "All evaluated rows retain compatibility, feasibility, rejection reason, cost, energy, terminal SOE, unmet energy, and estimated-mass fields.",
    "The Optimization Ranking tab receives the top feasible table, limited to ten rows.",
    "The best feasible result becomes the app's current result, so the Signals, Detailed Plot, and KPI tabs update to that configuration.",
    "Cancel requests are honored between candidate simulations; the already completed rows remain the meaningful partial search history in memory.",
]))

section(story, "12.8", "Worked ranking and responsible interpretation")
best_config = top_configs.iloc[0]
story.append(P(
    f"The refreshed 144-case Battery 1 / Motor sweep retains the ten highest-ranked feasible rows. "
    f"The leading row uses {best_config['Battery1ID']} with {best_config['MotorID']} at approximately "
    f"{best_config['CostPer_km']:.4f} EUR/km under the auxiliary-active-dump regeneration policy. "
    "Because all component values are synthetic, this ranking demonstrates workflow rather than a procurement recommendation.",
    "BodyTextBook"))
opt_rows = [["Rank", "Battery 1", "Motor", "Cost (EUR/km)", "Electricity (kWh/km)"]]
for i, row in top_configs.head(5).iterrows():
    opt_rows.append([i + 1, row["Battery1ID"], row["MotorID"], f"{row['CostPer_km']:.4f}", f"{row['Electrical_kWh_per_km']:.4f}"])
story += [table_caption("12.3", "Top five feasible rows in the current 144-case study"), make_table(opt_rows, [16*mm, 34*mm, 34*mm, 43*mm, 43*mm], alignments=["CENTER", "CENTER", "CENTER", "RIGHT", "RIGHT"])]
release_audit_rows = [
    ["App-shaped release audit", "Evaluated", "Feasible", "Top sorted", "Best returned"],
    ["Hybrid default", "40", "4", "Yes", "Yes"],
    ["BEV, 1.0 set / two packs", "40", "2", "Yes", "Yes; fuel = 0"],
    ["BEV, 0.5 set / one pack", "1 smoke case", "1", "Yes", "Yes; Battery 2 count = 0"],
]
story += [table_caption("12.4", "Post-fix Optimize verification using the app's actual configuration handoff"), make_table(release_audit_rows, [58*mm, 25*mm, 25*mm, 30*mm, 32*mm], font_size=6.8, alignments=["LEFT","CENTER","CENTER","CENTER","CENTER"])]
story.append(callout("Correct interpretation", "A first-ranked row is the best feasible row within the combinations that were actually evaluated, under the selected route, mass, prices, initial SOEs, component data, and terminal-energy rule. It is not automatically a production recommendation or a global optimum when the evaluation budget truncates the enabled search space.", "warn"))

# Chapter 13
chapter(story, "13", "Route Library and Provenance")
section(story, "13.1", "Route categories")
story.append(P(f"The database contains {len(route_catalog)} routes: three official EU VECTO passenger missions, eight geographic long-distance coach corridors, and nine geographic German city cycles for Mannheim, Stuttgart, Berlin, Munich, Hamburg, Frankfurt, Cologne, Dusseldorf, and Leipzig. The former synthetic concept routes have been removed. Route_Catalog records region, source basis, organization, URL, retrieval time, licence, distance, duration, maximum speed, geolocation availability, endpoint coordinates, and notes.", "BodyTextBook"))
section(story, "13.2", "Geographic route geometry and terrain")
story.append(P("Route_Geometry stores 20,257 ordered latitude/longitude samples for seventeen geographic routes, together with cumulative polyline distance, geometry source, retrieval timestamp, cached elevation, and elevation provenance. Coordinates come from full GeoJSON route overviews returned by OSRM over OpenStreetMap road data. Terrain height is added as a separate Copernicus DEM GLO-90 channel for the 3D Route Map. The app can display either elevation in metres or derived slope in percent without removing the original 2D map. Elevation-derived slope is visualization evidence only: it is not silently substituted into the zero-grade long-route simulation channel.", "BodyTextBook"))
story.append(P("Each German city circuit uses project-defined urban waypoints snapped to the driving network, a 50 km/h cap, 1.0 m/s2 acceleration and 1.3 m/s2 braking limits, and 20-second dwell events at nominal 800 m passenger-stop spacing. VECTO declaration cycles are not tied to a unique real road, so coordinates are intentionally not fabricated for them.", "BodyTextBook"))
story += [image_flow(ASSET / "app_mannheim_city_map.png", max_width=170*mm, max_height=102*mm), figure_caption("13.1", "Mannheim geographic city circuit selected in the Route Map tab")]
section(story, "13.3", "VECTO missions")
story.append(P("The Urban, Suburban, and Coach declaration-mode mission cycles are unmodified source files from the European Commission VECTO repository. The converter applies a deterministic bus driver with 1.0 m/s2 acceleration and 1.3 m/s2 braking bounds, honors declared stops, preserves the raw distance-domain trace, and produces a 1 Hz time history for this model.", "BodyTextBook"))
urban_fig = line_chart(urban["Time_s"] / 60, [("Speed", urban["Speed_kmh"], BLUE)], "EU VECTO Urban speed profile", "Time (min)", "Speed (km/h)", y_min=0)
story += [urban_fig, figure_caption("13.2", "Default EU VECTO Urban mission after deterministic time conversion")]
section(story, "13.4", "Long European corridors")
long_routes = route_catalog[route_catalog["RouteType"] == "Long-distance coach"].copy()
route_labels = [str(r).replace("EUR-", "") for r in long_routes["RouteID"]]
long_fig = horizontal_bar_chart(route_labels, list(long_routes["Distance_km"]), "OSM/OSRM long-distance route library", " km", max_value=1000)
story += [long_fig, figure_caption("13.3", "Geographic coach corridors in the requested 600-1000 km band")]
long_rows = [["Route", "Countries", "Distance", "Adapted duration"]]
for _, row in long_routes.iterrows():
    long_rows.append([row["RouteName"], row["Region"], f"{row['Distance_km']:.1f} km", f"{row['Duration_s']/3600:.2f} h"])
story += [table_caption("13.1", "Long-route source distances and adapted mission durations"), make_table(long_rows, [67*mm, 51*mm, 24*mm, 28*mm], font_size=6.7, alignments=["LEFT", "LEFT", "RIGHT", "RIGHT"])]
story.append(callout("Route fidelity warning", "OSRM supplies road-network distance and estimated segment durations but not elevation. Copernicus DEM GLO-90 now supports terrain visualization, but long-route simulation grade remains zero until a smoothed, road-aligned grade trace is derived and validated. Mountain-route energy is therefore still understated.", "warn"))

# Chapter 14
chapter(story, "14", "Verification, Validation, and Test Evidence")
section(story, "14.1", "Layered evidence")
story += [validation_pyramid(), figure_caption("14.1", "Verification pyramid used by the automated suite")]
story.append(P(f"The automated behavioral suite combines analytical comparisons, traction isolation, constant genset power, regeneration priority, pneumatic brake blending, modular data discovery, dynamic battery maps, motor loss maps, saturation, conservation, repeatability, optimizer ranking, scalable battery sets, calculated mass, constrained-speed behavior, depletion, and forward performance. All {len(tests)} behavioral scenarios currently pass. The complete MATLAB unit and integration suite contains 56 passing tests, including Hybrid/BEV implementation equivalence and component-extension coverage.", "BodyTextBook"))
section(story, "14.2", "Test matrix")
test_rows = [["Test", "Purpose", "Result"]]
for _, row in tests.iterrows():
    test_rows.append([row["Test"], row["Purpose"], row["Status"]])
story += [table_caption("14.1", "Automated MATLAB test suite"), make_table(test_rows, [47*mm, 101*mm, 22*mm], font_size=6.6, alignments=["LEFT", "LEFT", "CENTER"])]
section(story, "14.3", "Independent hand check")
story.append(P("For the current calculated 18,763.636 kg default mass, the constant-speed level-road test compares model wheel power against an independent analytical expression. Both calculations give 20.542 kW at the test condition. The test also confirms that 15-tonne base curb, installed battery mass, genset mass, and entered load propagate into the longitudinal dynamics.", "BodyTextBook"))
story.append(callout("Passing tests do not validate synthetic data", "Tests show that the implementation behaves consistently with its equations and stated limits. They do not prove that catalog efficiencies, BSFC values, component ratings, or route assumptions represent a production vehicle.", "critical"))
section(story, "14.4", "Requirements-to-evidence traceability")
story.append(P("The release evidence package links each concept requirement to its equation or architectural rule, observable signal, verification scenario, and numerical acceptance criterion. This closes the argument from claim to evidence and makes missing proof visible. The complete machine-readable matrix is stored in Requirements_Traceability_Matrix.csv and the human-readable version is stored in documentation/Requirements_Traceability_Matrix.md.", "BodyTextBook"))
gate_rows=[["Credibility gate","Status","Decision"]]
for _, row in credibility_gates.iterrows():
    gate_rows.append([row["Gate"],row["Status"],row["Decision"]])
story += [table_caption("14.2", "Current model-credibility gates"), make_table(gate_rows,[54*mm,32*mm,84*mm],font_size=6.5)]
story.append(callout("Gate discipline", "A concept-verification PASS does not imply implementation-equivalence PASS, supplier calibration, or measured-vehicle validation. Management decisions must use the lowest unresolved gate relevant to the decision being made.", "critical"))
section(story, "14.5", "Independent implementation equivalence")
story.append(P("The MATLAB kernels and Simulink models are independent implementations fed from the same selected configuration. Separate Hybrid and BEV cases compare speed, wheel demand, motor DC power, pneumatic friction braking, auxiliaries, both battery SOEs, genset power, total fuel, and integrated balance residual against declared tolerances. All twenty checks currently pass. This demonstrates numerical implementation equivalence for the declared cases; it is not supplier calibration or measured-vehicle validation.", "BodyTextBook"))
eq_rows=[["Mode","Signal","Max error","Tolerance","Unit","Status"]]
for _, row in equivalence_checks.iterrows():
    eq_rows.append([row["Powertrain"],row["Signal"],f'{row["MaxAbsoluteError"]:.4g}',f'{row["Tolerance"]:.4g}',row["Unit"],row["Status"]])
story += [table_caption("14.3", "Hybrid and BEV MATLAB-Simulink signal-level equivalence"), make_table(eq_rows,[20*mm,48*mm,25*mm,24*mm,25*mm,28*mm],font_size=6.2)]
section(story, "14.6", "Decision robustness and concept baselines")
story.append(P("The sensitivity study perturbs mass, aerodynamic drag, rolling resistance, auxiliary demand, motor efficiency, and battery discharge efficiency one factor at a time. It is deterministic screening, not a probability distribution. The concept comparison places the implemented hybrid beside analytical battery-electric and conventional-diesel screening baselines. Those baselines state their efficiency and fuel-energy assumptions and are suitable for architecture discussion, not procurement or homologation.", "BodyTextBook"))
concept_rows=[["Concept","Cost (EUR/km)","Source energy (kWh/km)","Evidence"]]
for _, row in concept_comparison.iterrows():
    concept_rows.append([row["Concept"],f'{row["Cost_EUR_per_km"]:.3f}',f'{row["SourceEnergy_kWh_per_km"]:.3f}',row["EvidenceLevel"]])
story += [table_caption("14.4", "Transparent powertrain concept screening"), make_table(concept_rows,[72*mm,31*mm,38*mm,29*mm],font_size=6.5)]

# Chapter 15
chapter(story, "15", "Worked Example: Default VECTO Urban Case")
section(story, "15.1", "Case definition")
story.append(P(f"The default Hybrid case uses the VECTO Urban route, one battery set containing two BAT-12 packs, two MOT-12 rear hub motors, GEN-05, {summary['CalculatedCurbMass_kg']/1000:.3f} tonnes calculated curb mass, {summary['LoadMass_t']:.1f} tonnes entered load, and {summary['EstimatedVehicleMass_kg']/1000:.3f} tonnes total mass. Initial Battery 1/2 SOEs are {100*float(dashboard.get('InitialBattery1SOE')):.0f}% and {100*float(dashboard.get('InitialBattery2SOE')):.0f}%; energy prices are {float(dashboard.get('FuelPrice')):.2f} EUR/L diesel and {float(dashboard.get('ElectricityPrice')):.2f} EUR/kWh electricity.", "BodyTextBook"))
section(story, "15.2", "Signal interpretation")
sig_fig = line_chart(
    signals["Time_s"] / 60,
    [("B1 SOE", signals["Battery1SOE_pct"], BLUE), ("B2 SOE", signals["Battery2SOE_pct"], ORANGE)],
    "Battery state of energy on the default mission", "Time (min)", "SOE (%)", y_min=45, y_max=86,
)
story += [sig_fig, figure_caption("15.1", "Default dual-battery SOE histories and active-pack behavior")]
story.append(P(
    f"Battery 1 is initially active and ends at {100*summary['FinalBattery1SOE']:.1f}% SOE; "
    f"Battery 2 ends at {100*summary['FinalBattery2SOE']:.1f}% SOE. Regeneration is no longer redirected "
    f"to the standby pack: {summary['RegenerationToAuxiliaryEnergy_kWh']:.3f} kWh supplies auxiliaries first, "
    f"{summary['RegenerationToActiveBatteryEnergy_kWh']:.3f} kWh is accepted at the active-battery terminal, "
    f"and {summary['ResistorLoadBankEnergy_kWh']:.3f} kWh is dissipated in the resistor bank. "
    "The genset does not start because the standby pack remains above the fixed 30% start condition.",
    "BodyTextBook"))
section(story, "15.3", "Summary metrics")
summary_rows = [
    ["Metric", "Value", "Interpretation"],
    ["Route distance", f"{summary['RouteDistance_km']:.3f} km", "distance integrated from the prepared 1 s trace"],
    ["Grid-equivalent energy", f"{summary['GridEquivalentEnergy_kWh']:.3f} kWh", "terminal pack deficits divided by grid efficiency"],
    ["Electrical intensity", f"{summary['Electrical_kWh_per_km']:.4f} kWh/km", "equivalent replenishment basis"],
    ["Operating cost", f"{summary['CostPer_km']:.4f} EUR/km", "fuel plus grid electricity divided by distance"],
    ["Available regenerated energy", f"{summary['RegeneratedEnergy_kWh']:.3f} kWh", "non-negative energy available from negative motor DC power"],
    ["Regeneration to auxiliaries", f"{summary['RegenerationToAuxiliaryEnergy_kWh']:.3f} kWh", "first-priority direct auxiliary supply"],
    ["Regeneration to active battery", f"{summary['RegenerationToActiveBatteryEnergy_kWh']:.3f} kWh", "second-priority accepted terminal charge energy"],
    ["Resistor-bank energy", f"{summary['ResistorLoadBankEnergy_kWh']:.3f} kWh", "third-priority controlled waste heat"],
    ["Auxiliary energy", f"{summary['AuxiliaryEnergy_kWh']:.3f} kWh", "base and HVAC electrical demand"],
    ["Unmet traction energy", f"{summary['UnmetTractionEnergy_kWh']:.6f} kWh", "zero: case is feasible under the project criterion"],
    ["Energy-balance error", f"{summary['EnergyBalanceError_kWh']:.9f} kWh", "zero within numerical reporting precision"],
]
story += [table_caption("15.1", "Default-case results"), make_table(summary_rows, [51*mm, 40*mm, 79*mm], alignments=["LEFT", "RIGHT", "LEFT"])]
story.append(callout("Worked-example conclusion", "The case is internally feasible and conservative under equivalent replenishment. It is not charge sustaining because combined terminal SOE is below the initial value, and it is not a real-vehicle performance claim.", "info"))

# Chapter 16
chapter(story, "16", "Long-Route and Mass Feasibility Studies")
section(story, "16.1", "Mass sweep")
mass_cost_fig = line_chart(
    mass_sweep["TotalVehicleMass_kg"] / 1000,
    [("Cost", mass_sweep["CostPer_km"], TEAL)],
    "Reported operating cost versus total mass", "Total vehicle mass (tonnes)", "EUR/km", y_min=0.3,
)
story += [mass_cost_fig, figure_caption("16.1", "Cost trend must be read together with unmet traction energy")]
story.append(P(f"At {mass_sweep.iloc[0]['TotalVehicleMass_kg']/1000:.0f} tonnes the default route has {mass_sweep.iloc[0]['UnmetEnergy_kWh']:.3f} kWh unmet traction energy. The entered payload is increased at each study point while the selected hardware-dependent curb mass remains fixed; unmet energy reaches {mass_sweep.iloc[-1]['UnmetEnergy_kWh']:.3f} kWh at {mass_sweep.iloc[-1]['TotalVehicleMass_kg']/1000:.0f} tonnes. A cost trend alone therefore mixes feasible and infeasible cases; the feasibility flag is governing.", "BodyTextBook"))
section(story, "16.2", "European long-route study")
unmet_fig = horizontal_bar_chart(
    [x.replace("EUR-", "") for x in long_study["RouteID"]],
    list(long_study["UnmetEnergy_kWh"]),
    f"Unmet energy with the default {summary['EstimatedVehicleMass_kg']/1000:.3f}-tonne component set", " kWh", max_value=1200,
    colors_list=[RED] * len(long_study),
)
story += [unmet_fig, figure_caption("16.2", "All long-route default cases are infeasible despite apparently low reported costs")]
story.append(P(f"The long missions integrate correctly, but the current default architecture cannot supply the complete requested traction/DC energy after stored energy is depleted. Unmet energy ranges from {long_study['UnmetEnergy_kWh'].min():.0f} to {long_study['UnmetEnergy_kWh'].max():.0f} kWh. The low cost values in these runs are therefore not design candidates; they are examples of why feasibility must precede economics.", "Lead"))
long_result_rows = [["Route", "Simulated km", "Fuel (L)", "Unmet (kWh)", "Feasible?"]]
for _, row in long_study.iterrows():
    long_result_rows.append([
        row["RouteID"], f"{row['SimDistance_km']:.1f}", f"{row['Fuel_L']:.1f}",
        f"{row['UnmetEnergy_kWh']:.1f}", "No",
    ])
story += [table_caption("16.1", "Default-component long-route outcomes"), make_table(long_result_rows, [39*mm, 32*mm, 30*mm, 38*mm, 31*mm], alignments=["LEFT", "RIGHT", "RIGHT", "RIGHT", "CENTER"])]
section(story, "16.3", "How to make a long-route study meaningful")
story.append(numbered_list([
    "Increase motor and active-pack capability, add sufficient battery energy, or schedule external charging; do not count genset output as traction power.",
    "Introduce realistic elevation and road-grade data before comparing Alpine corridors.",
    "Calibrate speed, dwell, passenger load, HVAC, traffic, and rest-stop assumptions with measured or operator data.",
    "Run bounded optimization with feasibility constraints, then compare equivalent replenishment and charge-sustaining results.",
    "Repeat under temperature, mass, headwind, and component-degradation scenarios.",
]))

# Chapter 17
chapter(story, "17", "Limitations, Responsible Use, and Extensions")
section(story, "17.1", "Current model boundary")
limitations = [
    ["Missing fidelity", "Why it matters", "Recommended extension"],
    ["Battery dynamic temperature and ageing", "Current maps use scenario temperature but do not integrate cell temperature or degradation", "Calibrated electrothermal states plus cycle/calendar ageing"],
    ["Ageing", "Usable energy, resistance, and cost evolve with throughput and temperature", "Cycle/calendar degradation and life-cycle cost"],
    ["Driver and plant calibration", "Forward and constrained formulations are implemented but use concept calibrations", "Calibrate driver gains, force limits, stall thresholds, and plant response with vehicle data"],
    ["Validated simulation grade for long routes", "Cached terrain supports 3D display, but zero dynamics grade still understates mountain energy", "Map-match, smooth, and validate an elevation-derived grade channel"],
    ["Engine transients and emissions", "Fuel and pollutants depend on warm-up and aftertreatment", "Dynamic engine/aftertreatment model and measured maps"],
    ["Wheel slip and lateral dynamics", "Traction limits and stability are not represented", "Tyre-road force limits and vehicle dynamics"],
    ["Supplier data", "Synthetic catalogs cannot support procurement", "Version-controlled, validated manufacturer maps"],
]
story += [table_caption("17.1", "Fidelity gaps and extension paths"), make_table(limitations, [45*mm, 60*mm, 65*mm], font_size=6.7)]
section(story, "17.2", "Recommended development roadmap")
story += [roadmap_diagram(), figure_caption("17.1", "A disciplined progression from concept model to validated engineering platform")]
story.append(PageBreak())
section(story, "17.3", "Graduate exercises")
story.append(numbered_list([
    "Derive an analytical constant-speed energy-per-kilometre expression and compare it with the kernel over several masses and headwinds.",
    "Replace scalar motor efficiency with a two-dimensional torque-speed map and quantify the change in route energy.",
    "Add elevation to one OSM/OSRM route, validate grade smoothing, and compare energy with the zero-grade baseline.",
    "Design an alternative supervisor that allows controlled parallel discharge and define new safety/feasibility tests.",
    "Compare 0.5-, 1.0-, 1.5-, and 2.0-set BEV operation, including capability sharing, route completion, terminal SOE, and mass.",
    "Formulate a multi-objective search for cost, fuel, battery throughput, and component mass; produce a Pareto front.",
    "Create a charge-sustaining urban case by sizing battery energy and the fixed-point standby charger; keep the 30% role threshold unchanged.",
    "Add a simple battery thermal state and demonstrate a cold-weather power-derating scenario.",
    "Construct a requirements-to-test traceability matrix for every limit and KPI in the model.",
]))
story.append(callout("Final engineering lesson", "A useful model is not only a set of equations. It is an agreement among data definitions, physical assumptions, interfaces, numerical methods, controller logic, evidence, and interpretation. When any one of those changes, the others must be reviewed.", "note"))

# Appendix A
chapter(story, "Appendix A", "Quick-Start Laboratory")
story.append(P("Use the following sequence for a first supervised laboratory session.", "Lead"))
story.append(numbered_list([
    "Open MATLAB R2025a and change to the project folder.",
    "Run startup_hybrid_bus and confirm that the project loads without startup issues.",
    "Run Results = run_hybrid_bus_simulation; inspect Results.Summary and Results.Validation.",
    "Open HybridBus_BackwardModel.slx and trace signals from route input to energy accounting.",
    "Launch app = HybridBusApp; repeat the default case and inspect all plot tabs.",
    "Change only one factor, such as entered load or route, and explain every KPI change before running a second factor.",
    "Run TestResults = run_all_hybrid_bus_tests; confirm 50 PASS behavioral scenarios, then run the full tests folder and confirm 56 passing unit/integration tests.",
    "Run Credibility = generate_model_credibility_report; review every PASS, FAIL, NOT AVAILABLE, and NOT IMPLEMENTED gate before presenting a decision.",
]))
story.append(code_block("""
db = load_hybrid_bus_database(fullfile("data","HybridBus_ComponentDatabase.xlsx"));
report = validate_hybrid_bus_database(db);
assert(report.IsValid, strjoin(report.Errors,newline));

overrides = struct('SelectedRoute',"EUR-BER-VIE", ...
                   'PowertrainMode',"Hybrid", ...
                   'BatterySetMultiplier',1, ...
                   'LoadMass_t',5);
Results = run_hybrid_bus_simulation( ...
    fullfile("data","HybridBus_ComponentDatabase.xlsx"),overrides,'SaveResults',false);
Sequence = run_powertrain_sequence(db,overrides,true);
"""))

# Appendix B
chapter(story, "Appendix B", "Signal and Parameter Reference")
signal_rows = [
    ["Group", "Signal", "Unit", "Meaning"],
    ["Vehicle", "Speed_m_s / Acceleration_m_s2", "m/s, m/s2", "prescribed speed and filtered acceleration"],
    ["Wheel", "Demand_kW / Delivered_kW", "kW", "required and limit-delivered wheel power"],
    ["Wheel", "UnmetTraction_kW / UnmetRegen_kW", "kW", "mechanical shortfalls"],
    ["Motors", "ElectricalPower_kW / MotorSpeed_rpm", "kW, rpm", "DC demand/regeneration and motor speed"],
    ["Auxiliary", "Power_kW", "kW", "base, HVAC, and route-multiplied load"],
    ["Regeneration", "Available_kW / ToAuxiliary_kW", "kW", "available braking energy and first-priority allocation"],
    ["Regeneration", "ToActiveBattery_kW / ResistorLoadBank_kW", "kW", "second- and third-priority allocations"],
    ["Battery 1/2", "Power_kW / Energy_kWh / SOE", "kW, kWh, 1", "pack source/sink and stored state"],
    ["Genset", "ElectricalPower_kW / FuelRate_L_s", "kW, L/s", "constant standby-charger power and fuel flow"],
    ["Genset", "ChargeDestinationBattery", "integer", "standby pack receiving genset power; zero when off"],
    ["Controller", "ActiveBattery / StandbyBattery / Mode", "integer", "supervisory role and discrete mode"],
    ["Controller", "ConnectedBatteryCount", "integer", "active-bank count in Hybrid; total connected pack count in BEV"],
    ["Energy", "BalanceResidual_kW", "kW", "power-conservation diagnostic"],
    ["Energy", "UnmetDCPower_kW / RejectedCharge_kW", "kW", "source/sink saturation accounting including load-bank power"],
    ["Energy", "RejectedGensetCharge_kW", "kW", "charger mismatch kept distinct from regenerative dissipation"],
]
story += [table_caption("B.1", "Principal logged signals"), make_table(signal_rows, [26*mm, 57*mm, 28*mm, 59*mm], font_size=6.6)]
param_rows = [
    ["Area", "Principal parameters", "Workbook source"],
    ["Vehicle", "15 t base curb, installed battery/genset mass, entered load, gravity, Cd, frontal area", "Calculated / Battery / Genset / Dashboard"],
    ["Route", "time, speed, grade, dwell, auxiliary multiplier", "Route sheets"],
    ["Hub drive", "wheel radius, ratio, efficiencies, torque/power/speed", "Tyre / Final Drive / Motor"],
    ["Battery", "usable energy, SOE bounds, efficiencies, power, regen, set multiplier, deterministic pack counts", "Battery + Dashboard/app"],
    ["Genset", "optimum/max power, minimum times, BSFC, efficiency", "Genset + maps"],
    ["Supervisor", "Hybrid 30% role swap or BEV parallel capability sharing; auxiliary-battery-dump regen priority", "Architecture mode + control calibration"],
    ["Load bank", "regenerative surplus after active-pack limits", "Derived power sink"],
    ["Economics", "fuel tank, prices, grid efficiency, terminal method", "Dashboard / Vehicle / Prices / Optimization"],
]
story += [table_caption("B.2", "Principal parameter groups"), make_table(param_rows, [34*mm, 88*mm, 48*mm], font_size=6.8)]

# References
chapter(story, "References", "Sources and Project Files")
refs = [
    f"[1] HybridBusProject source tree, modular route/component data, MATLAB functions, Hybrid, BEV, and Performance Simulink models, generated results, traceability matrix, and model-credibility evidence. Database version {dashboard.get('DatabaseVersion')}; textbook revision 2.3, revised 31 August 2026.",
    "[2] European Commission, Vehicle Energy Consumption Calculation Tool (VECTO), https://climate.ec.europa.eu/eu-action/transport-decarbonisation/road-transport/vehicle-energy-consumption-calculation-tool-vecto_en.",
    "[3] European Commission VECTO source repository, declaration mission cycles, commit d16ba783c1d8af0ea68797f5d7ed6cf01877d402, https://code.europa.eu/vecto/vecto.",
    "[4] Project OSRM API documentation, route service and annotation objects, https://project-osrm.org/docs/v5.24.0/api/.",
    "[5] OpenStreetMap contributors, Copyright and Open Database License attribution, https://www.openstreetmap.org/copyright.",
    "[6] European Commission Mobility and Transport, Driving time and rest periods, including a 45-minute break after 4.5 hours of driving, https://transport.ec.europa.eu/transport-modes/road/social-provisions/driving-time-and-rest-periods_en.",
    "[7] Project documentation: System_Architecture.md, Mode_Transition_Table.md, Parameter_Dictionary.md, Signal_Dictionary.md, Assumptions_and_Limitations.md, Test_Report.md, Model_Credibility_Report.md, Requirements_Traceability_Matrix.md, and Validation_Data_and_Extension_Plan.md.",
]
for ref in refs:
    story.append(P(escape(ref), "Reference"))


doc = TextbookDocTemplate(
    str(OUTPUT), pagesize=A4, rightMargin=19*mm, leftMargin=19*mm,
    topMargin=17*mm, bottomMargin=15*mm,
    title="Hybrid-Electric Bus System Modeling",
    author="HybridBusProject",
    subject="Graduate textbook for MATLAB/Simulink hybrid-electric bus energy modeling",
)
doc.multiBuild(story)
print(f"CREATED {OUTPUT}")
